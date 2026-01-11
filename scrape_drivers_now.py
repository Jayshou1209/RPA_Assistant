"""
快速爬取司机数据并生成Excel
"""
import os
import sys
from datetime import datetime
from api_client import APIClient
from real_api_scraper import RealAPIScraper
import config

def main():
    print("="*60)
    print("开始爬取司机完整数据...")
    print("="*60)
    
    # 读取Token
    try:
        with open("token.txt", 'r', encoding='utf-8-sig') as f:
            token = f.read().strip()
        print(f"✓ Token已加载 (长度: {len(token)} 字符)")
    except Exception as e:
        print(f"✗ 读取Token失败: {e}")
        return
    
    # 初始化客户端
    try:
        api_client = APIClient(token)
        scraper = RealAPIScraper(api_client)
        print("✓ API客户端初始化成功")
    except Exception as e:
        print(f"✗ 初始化失败: {e}")
        return
    
    # 测试连接
    print("\n测试API连接...")
    success, message = api_client.verify_connection()
    if success:
        print(f"✓ {message}")
    else:
        print(f"✗ 连接失败: {message}")
        return
    
    # 开始爬取
    print("\n" + "="*60)
    print("开始爬取司机数据（包含完整证件和车辆信息）...")
    print("="*60)
    
    def progress_callback(current, total, msg):
        print(f"  [{current}/{total}] {msg}")
    
    try:
        drivers = scraper.get_all_drivers_with_full_details(
            per_page=100,
            progress_callback=progress_callback
        )
        
        print("\n" + "="*60)
        print(f"✓ 成功爬取 {len(drivers)} 位司机的数据")
        
        # 过滤只保留Active状态的司机
        active_drivers = []
        for driver in drivers:
            driver_data = driver.get('driver', driver)
            status = driver_data.get('status', '')
            if status == 'active':
                active_drivers.append(driver)
        
        print(f"✓ 过滤后 Active 状态司机: {len(active_drivers)} 位")
        print("="*60)
        
        # 使用过滤后的司机数据
        drivers = active_drivers
        
        # 导出Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = os.path.join(config.DATA_DIR, f"司机完整数据_Active_{timestamp}.xlsx")
        
        print(f"\n开始导出Excel文件...")
        print(f"文件路径: {excel_file}")
        
        # 调用导出函数
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "司机数据"
        
        # 定义表头
        headers = [
            "First Name", "Last Name", "Gender", "DOB", "Phone",
            "Social Security No", "License Number", "License State",
            "License Class", "License Issued Date", "License Expiration Date",
            "TLC License Number", "TLC License State", "TLC License Expiration Date",
            "Commercial CDL License Number", "Commercial CDL License State",
            "Commercial CDL License Effective Date", "Commercial CDL License Expiration Date",
            "Background Check Issue Date", "Background Check Expiration Date",
            "Drug Test Issue Date", "Drug Test Expiration Date",
            "MVR Issue Date", "MVR Expiration Date",
            "Clearinghouse Issue Date", "Clearinghouse Expiration Date",
            "Medical Examiner Issue Date", "Medical Examiner Expiration Date",
            "Ophthalmology Issue Date", "Ophthalmology Expiration Date",
            "Vehicle Model", "Vehicle Year", "Vehicle Color", "Vehicle License Plate",
            "Vehicle VIN", "FHV Diamond Number", "FHV Diamond Expiration Date",
            "Insurance Policy Number", "Insurance Company", "Insurance Expiration Date",
            "Registration Expiration Date", "NYS Inspection Sticker Expiration Date"
        ]
        
        # 写入表头
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_idx, driver in enumerate(drivers, start=2):
            # 获取driver对象（可能在顶层，也可能在driver字段中）
            driver_data = driver.get('driver', driver)
            
            # 基本信息
            ws.cell(row=row_idx, column=1, value=driver_data.get('first_name', ''))
            ws.cell(row=row_idx, column=2, value=driver_data.get('last_name', ''))
            ws.cell(row=row_idx, column=3, value=driver_data.get('sex', driver_data.get('gender', '')))
            ws.cell(row=row_idx, column=4, value=driver_data.get('dob_date', driver_data.get('dob', '')))
            ws.cell(row=row_idx, column=5, value=driver_data.get('phone_number', ''))
            
            # SSN
            ws.cell(row=row_idx, column=6, value=driver_data.get('ssn', ''))
            
            # Driver License (从driver_data直接获取)
            ws.cell(row=row_idx, column=7, value=driver_data.get('driver_license', ''))
            ws.cell(row=row_idx, column=8, value=driver_data.get('driver_license_state', ''))
            
            # 证件信息 - 从documents数组中获取
            documents = driver.get('documents', [])
            
            # 建立证书类型到列的映射 (使用API返回的type字段)
            cert_mapping = {
                'driver_license': (7, 8, 9, 10, 11),  # License Number, State, Class, Issue Date, Expiry Date
                'tlc_license': (12, 13, None, None, 14),  # TLC Number, State, N/A, N/A, Expiry Date
                'cdl': (15, 16, 17, 18),  # CDL Number, State, Effective Date, Expiry Date
                'background_check': (19, None, None, 20),  # Issue Date, Expiry Date
                'drug_test': (21, None, None, 22),  # Issue Date, Expiry Date
                'sentry_drug_test': (21, None, None, 22),  # Issue Date, Expiry Date
                'arro_drug_test': (21, None, None, 22),  # Issue Date, Expiry Date
                'ctg_drug_test': (21, None, None, 22),  # Issue Date, Expiry Date
                'mvr': (23, None, None, 24),  # Issue Date, Expiry Date
                'clearinghouse': (25, None, None, 26),  # Issue Date, Expiry Date
                'medical_examiner': (27, None, None, 28),  # Issue Date, Expiry Date
                'ophthalmology': (29, None, None, 30)  # Issue Date, Expiry Date
            }
            
            for doc in documents:
                doc_type = doc.get('type', '')  # 直接使用type字段
                
                if doc_type in cert_mapping:
                    cols = cert_mapping[doc_type]
                    
                    # 获取options中的数据
                    issue_date = None
                    license_class = None
                    options = doc.get('options')
                    if options and isinstance(options, list):
                        for opt in options:
                            if opt.get('name') == 'issue_date':
                                issue_date = opt.get('value')
                            elif opt.get('name') == 'license_class':
                                license_class = opt.get('value')
                    
                    # 写入证书编号
                    if cols[0]:
                        ws.cell(row=row_idx, column=cols[0], value=doc.get('number', ''))
                    
                    # 写入州
                    if cols[1]:
                        ws.cell(row=row_idx, column=cols[1], value=doc.get('state', ''))
                    
                    # 对于driver_license，写入License Class
                    if doc_type == 'driver_license' and len(cols) > 4:
                        if cols[2]:  # License Class列
                            ws.cell(row=row_idx, column=cols[2], value=license_class or '')
                        if cols[3]:  # Issue Date列
                            ws.cell(row=row_idx, column=cols[3], value=issue_date or '')
                        if cols[4]:  # Expiry Date列
                            ws.cell(row=row_idx, column=cols[4], value=doc.get('expires_at', ''))
                    elif doc_type == 'tlc_license':
                        # TLC License: number(12), state(13), expiry(14)
                        if cols[4]:  # Expiry Date列
                            ws.cell(row=row_idx, column=cols[4], value=doc.get('expires_at', ''))
                    else:
                        # 其他证件类型
                        if cols[2]:  # Issue Date列
                            ws.cell(row=row_idx, column=cols[2], value=issue_date or '')
                        if cols[3]:  # Expiry Date列
                            ws.cell(row=row_idx, column=cols[3], value=doc.get('expires_at', ''))
            
            # 车辆信息 - 需要获取车辆详细信息
            cars = driver.get('cars', [])
            if cars:
                car_id = cars[0].get('id')
                if car_id:
                    try:
                        # 获取车辆详细信息
                        car_detail_response = api_client.get(f'/fleet/cars/{car_id}')
                        car_detail = car_detail_response.get('car', {})
                        
                        # 基本信息：Model(31), Year(32), Color(33), Plate(34), VIN(35)
                        ws.cell(row=row_idx, column=31, value=car_detail.get('model', '').title())
                        ws.cell(row=row_idx, column=32, value=car_detail.get('year', ''))
                        ws.cell(row=row_idx, column=33, value=car_detail.get('color', '').title())
                        ws.cell(row=row_idx, column=34, value=car_detail.get('plate_number', ''))
                        ws.cell(row=row_idx, column=35, value=car_detail.get('vin_number', ''))
                        
                        # 从车辆documents中获取证件信息
                        car_docs = car_detail.get('documents', [])
                        for doc in car_docs:
                            doc_type = doc.get('type', '')
                            
                            if doc_type == 'fhv_diamond':
                                # FHV Diamond: Number(36), Expiration(37)
                                ws.cell(row=row_idx, column=36, value=doc.get('number', ''))
                                ws.cell(row=row_idx, column=37, value=doc.get('expires_at', ''))
                            
                            elif doc_type == 'insurance_id_card':
                                # Insurance: Policy Number(38), Company(39), Expiration(40)
                                ws.cell(row=row_idx, column=38, value=doc.get('number', ''))
                                
                                # 从options中获取保险公司
                                options = doc.get('options')
                                if options and isinstance(options, list):
                                    for opt in options:
                                        if opt.get('name') == 'insurance_company':
                                            ws.cell(row=row_idx, column=39, value=opt.get('value', ''))
                                            break
                                
                                ws.cell(row=row_idx, column=40, value=doc.get('expires_at', ''))
                            
                            elif doc_type == 'registration':
                                # Registration: Expiration(41)
                                ws.cell(row=row_idx, column=41, value=doc.get('expires_at', ''))
                            
                            elif doc_type == 'nys_inspection_sticker':
                                # NYS Inspection Sticker: Expiration(42)
                                ws.cell(row=row_idx, column=42, value=doc.get('expires_at', ''))
                                
                    except Exception as e:
                        print(f"  获取车辆 {car_id} 详情失败: {e}")
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = 20
        
        # 保存文件
        wb.save(excel_file)
        
        print(f"✓ Excel文件已生成: {excel_file}")
        print(f"✓ 总共导出 {len(drivers)} 条记录")
        
        # 显示文件信息
        file_size = os.path.getsize(excel_file)
        print(f"\n文件信息:")
        print(f"  - 文件大小: {file_size / 1024:.2f} KB")
        print(f"  - 记录数: {len(drivers)}")
        print(f"  - 列数: {len(headers)}")
        
        # 自动打开文件
        print(f"\n正在打开文件...")
        os.startfile(excel_file)
        
    except Exception as e:
        print(f"✗ 爬取失败: {e}")
        import traceback
        traceback.print_exc()
        return

if __name__ == "__main__":
    main()
    print("\n按任意键退出...")
    input()
