"""
RPA调度系统 - GUI可视化界面
集成所有功能：Token管理、数据爬取、派工、退工、订单转派
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import threading
import json
import os
from datetime import datetime, timedelta
from api_client import APIClient
from scraper import DataScraper
from dispatcher import Dispatcher
import config
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(config.LOG_FILE, encoding='utf-8'),
    ]
)
logger = logging.getLogger(__name__)


class RPAAutomationGUI:
    """RPA自动化系统GUI主界面"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("RPA调度系统自动化助手 v1.0")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)
        
        # 初始化变量
        self.api_client = None
        self.scraper = None
        self.dispatcher = None
        self.token_var = tk.StringVar(value=config.BEARER_TOKEN)
        self.status_var = tk.StringVar(value="就绪")
        self.last_data = None
        
        # 创建界面
        self.create_widgets()
        
        # 初始化API客户端
        self.initialize_client()
    
    def create_widgets(self):
        """创建界面组件"""
        # 主容器
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # 1. Token管理区域
        self.create_token_section(main_frame)
        
        # 2. 功能按钮区域
        self.create_function_buttons(main_frame)
        
        # 3. 输出显示区域
        self.create_output_section(main_frame)
        
        # 4. 状态栏
        self.create_status_bar(main_frame)
    
    def create_token_section(self, parent):
        """创建Token管理区域"""
        token_frame = ttk.LabelFrame(parent, text="📝 Token管理", padding="10")
        token_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Token输入
        ttk.Label(token_frame, text="Bearer Token:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        token_entry = ttk.Entry(token_frame, textvariable=self.token_var, width=80)
        token_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        
        # 按钮组
        btn_frame = ttk.Frame(token_frame)
        btn_frame.grid(row=0, column=2, padx=(5, 0))
        
        ttk.Button(btn_frame, text="💾 保存Token", command=self.save_token).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🔄 重新加载", command=self.reload_token).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="✓ 测试连接", command=self.test_connection).pack(side=tk.LEFT, padx=2)
        
        token_frame.columnconfigure(1, weight=1)
    
    def create_function_buttons(self, parent):
        """创建功能按钮区域"""
        # 左侧按钮面板
        btn_frame = ttk.Frame(parent)
        btn_frame.grid(row=1, column=0, sticky=(tk.N, tk.W, tk.E), padx=(0, 10))
        
        # 数据爬取
        data_frame = ttk.LabelFrame(btn_frame, text="📊 数据爬取", padding="10")
        data_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(data_frame, text="👤 爬取司机数据", command=self.scrape_drivers_only, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="📅 爬取排班数据", command=self.scrape_schedules_only, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="📦 爬取订单数据", command=self.scrape_orders_only, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="💰 生成账单", command=self.generate_billing, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="🚀 快速测试（10条）", command=self.quick_test_scrape, width=25).pack(fill=tk.X, pady=2)
        
        # 导出功能
        export_frame = ttk.LabelFrame(btn_frame, text="💾 导出与工具", padding="10")
        export_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(export_frame, text="📂 打开数据目录", command=self.open_data_folder, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(export_frame, text="🧹 清理数据目录", command=self.clean_data_folder, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(export_frame, text="📄 导出为JSON", command=self.export_json, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(export_frame, text="📊 导出为Excel", command=self.export_excel, width=25).pack(fill=tk.X, pady=2)
        
        # 调度操作
        dispatch_frame = ttk.LabelFrame(btn_frame, text="🎯 调度操作", padding="10")
        dispatch_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(dispatch_frame, text="➕ 派工 (Assign)", command=self.show_dispatch_dialog, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(dispatch_frame, text="🔄 转派 (Switch)", command=self.show_transfer_dialog, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(dispatch_frame, text="➖ 退工 (Revive)", command=self.show_withdraw_dialog, width=25).pack(fill=tk.X, pady=2)
        
        # 系统操作
        system_frame = ttk.LabelFrame(btn_frame, text="⚙️ 系统", padding="10")
        system_frame.pack(fill=tk.X)
        
        ttk.Button(system_frame, text="📜 查看日志文件", command=self.view_logs, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(system_frame, text="🗑️ 清空输出", command=self.clear_output, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(system_frame, text="ℹ️ 关于", command=self.show_about, width=25).pack(fill=tk.X, pady=2)
    
    def create_output_section(self, parent):
        """创建输出显示区域"""
        output_frame = ttk.LabelFrame(parent, text="📺 输出信息", padding="10")
        output_frame.grid(row=1, column=1, rowspan=2, sticky=(tk.N, tk.S, tk.E, tk.W))
        
        # 创建带滚动条的文本框
        self.output_text = scrolledtext.ScrolledText(
            output_frame, 
            wrap=tk.WORD, 
            width=80, 
            height=30,
            font=("Consolas", 10)
        )
        self.output_text.pack(fill=tk.BOTH, expand=True)
        
        # 配置标签样式
        self.output_text.tag_config("success", foreground="green", font=("Consolas", 10, "bold"))
        self.output_text.tag_config("error", foreground="red", font=("Consolas", 10, "bold"))
        self.output_text.tag_config("info", foreground="blue")
        self.output_text.tag_config("warning", foreground="orange")
        
        self.log("=" * 60)
        self.log("欢迎使用 RPA调度系统自动化助手", "info")
        self.log("=" * 60)
    
    def create_status_bar(self, parent):
        """创建状态栏"""
        status_frame = ttk.Frame(parent)
        status_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Label(status_frame, text="状态:").pack(side=tk.LEFT)
        status_label = ttk.Label(status_frame, textvariable=self.status_var, foreground="blue")
        status_label.pack(side=tk.LEFT, padx=5)
        
        # 添加时间显示
        self.time_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ttk.Label(status_frame, textvariable=self.time_var).pack(side=tk.RIGHT)
        self.update_time()
    
    def update_time(self):
        """更新时间显示"""
        self.time_var.set(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.root.after(1000, self.update_time)
    
    def log(self, message, tag=""):
        """输出日志"""
        self.output_text.insert(tk.END, f"{message}\n", tag)
        self.output_text.see(tk.END)
        self.root.update_idletasks()
    
    def set_status(self, status):
        """设置状态"""
        self.status_var.set(status)
        self.root.update_idletasks()
    
    def initialize_client(self):
        """初始化API客户端"""
        try:
            from enhanced_scraper import EnhancedScraper
            from real_api_scraper import RealAPIScraper
            self.api_client = APIClient(self.token_var.get())
            self.scraper = DataScraper(self.api_client)
            self.enhanced_scraper = EnhancedScraper(self.api_client)
            self.real_scraper = RealAPIScraper(self.api_client)
            self.dispatcher = Dispatcher(self.api_client)
            self.log("✓ API客户端初始化成功", "success")
        except Exception as e:
            self.log(f"✗ 初始化失败: {e}", "error")
    
    def save_token(self):
        """保存Token到配置文件"""
        try:
            new_token = self.token_var.get().strip()
            if not new_token:
                messagebox.showwarning("警告", "Token不能为空")
                return
            
            # 读取配置文件
            config_path = "config.py"
            with open(config_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 备份
            backup_path = f"config.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}.py"
            with open(backup_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            # 替换Token
            import re
            pattern = r'BEARER_TOKEN = "Bearer [^"]*"'
            replacement = f'BEARER_TOKEN = "{new_token}"'
            new_content = re.sub(pattern, replacement, content)
            
            # 更新日期
            date_pattern = r'# 最后更新时间: \d{4}-\d{2}-\d{2}'
            date_replacement = f"# 最后更新时间: {datetime.now().strftime('%Y-%m-%d')}"
            new_content = re.sub(date_pattern, date_replacement, new_content)
            
            # 保存
            with open(config_path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            
            # 重新初始化客户端
            config.BEARER_TOKEN = new_token
            self.initialize_client()
            
            self.log("✓ Token已保存并更新", "success")
            messagebox.showinfo("成功", f"Token已保存！\n备份文件: {backup_path}")
            
        except Exception as e:
            self.log(f"✗ 保存Token失败: {e}", "error")
            messagebox.showerror("错误", f"保存失败: {e}")
    
    def reload_token(self):
        """重新加载Token"""
        try:
            import importlib
            importlib.reload(config)
            self.token_var.set(config.BEARER_TOKEN)
            self.initialize_client()
            self.log("✓ Token已重新加载", "success")
        except Exception as e:
            self.log(f"✗ 重新加载失败: {e}", "error")
    
    def test_connection(self):
        """测试连接"""
        def test():
            self.set_status("测试连接中...")
            try:
                if self.api_client.verify_connection():
                    self.log("✓ 连接测试成功！", "success")
                    messagebox.showinfo("成功", "连接测试成功！")
                else:
                    self.log("✗ 连接测试失败", "error")
                    messagebox.showerror("失败", "连接测试失败，请检查Token")
            except Exception as e:
                self.log(f"✗ 连接错误: {e}", "error")
                messagebox.showerror("错误", f"连接错误: {e}")
            finally:
                self.set_status("就绪")
        
        threading.Thread(target=test, daemon=True).start()
    
    def scrape_all_data(self):
        """爬取所有数据"""
        def scrape():
            self.set_status("正在爬取数据...")
            self.log("\n" + "=" * 60)
            self.log("开始爬取所有数据...", "info")
            self.log("=" * 60)
            
            try:
                data = self.scraper.scrape_all_data()
                self.last_data = data
                
                # 保存数据
                self.scraper.save_data(data)
                
                self.log(f"\n✓ 数据爬取完成！", "success")
                self.log(f"  - 司机数量: {len(data.get('drivers', []))}")
                self.log(f"  - 车辆数量: {len(data.get('vehicles', []))}")
                self.log(f"  - 排班数量: {len(data.get('schedules', []))}")
                self.log(f"  - 数据已保存到: data/{config.DRIVER_DATA_FILE}")
                
                messagebox.showinfo("成功", "数据爬取完成！")
                
            except Exception as e:
                self.log(f"✗ 爬取失败: {e}", "error")
                messagebox.showerror("错误", f"爬取失败: {e}")
            finally:
                self.set_status("就绪")
        
        threading.Thread(target=scrape, daemon=True).start()
    
    def view_drivers(self):
        """查看司机列表"""
        def view():
            self.set_status("获取司机数据...")
            try:
                drivers = self.scraper.get_drivers()
                self.log("\n" + "=" * 60)
                self.log(f"司机列表 (共 {len(drivers)} 位)", "info")
                self.log("=" * 60)
                
                for i, driver in enumerate(drivers[:20], 1):
                    driver_info = f"{i}. ID:{driver.get('id')} | 姓名:{driver.get('name', '未知')} | 状态:{driver.get('status', '未知')}"
                    self.log(driver_info)
                
                if len(drivers) > 20:
                    self.log(f"\n... 还有 {len(drivers) - 20} 位司机")
                
            except Exception as e:
                self.log(f"✗ 获取失败: {e}", "error")
            finally:
                self.set_status("就绪")
        
        threading.Thread(target=view, daemon=True).start()
    
    def view_vehicles(self):
        """查看车辆列表"""
        def view():
            self.set_status("获取车辆数据...")
            try:
                vehicles = self.scraper.get_vehicles()
                self.log("\n" + "=" * 60)
                self.log(f"车辆列表 (共 {len(vehicles)} 辆)", "info")
                self.log("=" * 60)
                
                for i, vehicle in enumerate(vehicles[:20], 1):
                    vehicle_info = f"{i}. ID:{vehicle.get('id')} | 车牌:{vehicle.get('plate', '未知')} | 车型:{vehicle.get('type', '未知')}"
                    self.log(vehicle_info)
                
                if len(vehicles) > 20:
                    self.log(f"\n... 还有 {len(vehicles) - 20} 辆车")
                
            except Exception as e:
                self.log(f"✗ 获取失败: {e}", "error")
            finally:
                self.set_status("就绪")
        
        threading.Thread(target=view, daemon=True).start()
    
    def scrape_drivers_only(self):
        """只爬取司机数据（不含排班）"""
        def task():
            try:
                self.set_status("正在爬取司机数据...")
                self.log("=" * 60)
                self.log("开始爬取司机数据", "info")
                self.log("=" * 60)
                
                # 1. 获取司机基本信息
                self.log("\n1️⃣ 获取司机基本信息...", "info")
                drivers = self.real_scraper.get_all_drivers(per_page=100)
                self.log(f"✓ 获取到 {len(drivers)} 位司机基本信息", "success")
                
                # 2. 获取司机详细信息
                self.log("\n2️⃣ 获取司机详细资料...", "info")
                driver_details = []
                for i, driver in enumerate(drivers, 1):
                    driver_id = driver.get('id')
                    try:
                        detail = self.real_scraper.get_driver_detail(driver_id)
                        if detail:
                            driver_details.append({**driver, **detail})
                        else:
                            driver_details.append(driver)
                        if i % 10 == 0:
                            self.log(f"  已处理 {i}/{len(drivers)} 位司机...")
                    except Exception as e:
                        driver_details.append(driver)
                        self.log(f"  警告: 司机 {driver_id} 详情获取失败", "warning")
                
                self.log(f"✓ 获取到 {len(driver_details)} 位司机完整资料", "success")
                
                # 3. 导出Excel
                self.log("\n3️⃣ 导出Excel...", "info")
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_file = os.path.join(config.DATA_DIR, f'司机数据_{timestamp}.xlsx')
                
                import pandas as pd
                df = pd.DataFrame(driver_details)
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='司机信息', index=False)
                
                self.log(f"✓ Excel已保存: {excel_file}", "success")
                
                # 保存到last_data
                self.last_data = {'drivers': driver_details}
                
                self.log("\n" + "=" * 60)
                self.log("✓ 完成！", "success")
                self.log(f"司机总数: {len(driver_details)} 位", "info")
                self.log("=" * 60)
                
                self.set_status("就绪")
                messagebox.showinfo("完成", f"司机数据爬取完成！\n\n"
                                   f"总数: {len(driver_details)} 位\n"
                                   f"文件: {excel_file}")
                
            except Exception as e:
                import traceback
                self.log(f"✗ 爬取失败: {e}", "error")
                self.log(traceback.format_exc(), "error")
                self.set_status("就绪")
                messagebox.showerror("错误", f"爬取失败:\n{e}")
        
        threading.Thread(target=task, daemon=True).start()
    
    def scrape_schedules_only(self):
        """只爬取排班数据"""
        # 创建日期选择对话框
        date_dialog = tk.Toplevel(self.root)
        date_dialog.title("选择爬取日期")
        date_dialog.geometry("350x200")
        date_dialog.transient(self.root)
        date_dialog.grab_set()
        
        # 居中显示
        date_dialog.update_idletasks()
        x = (date_dialog.winfo_screenwidth() // 2) - (350 // 2)
        y = (date_dialog.winfo_screenheight() // 2) - (200 // 2)
        date_dialog.geometry(f"350x200+{x}+{y}")
        
        ttk.Label(date_dialog, text="选择要爬取排班数据的日期:", font=("Arial", 11)).pack(pady=15)
        
        # 日期输入框
        date_frame = ttk.Frame(date_dialog)
        date_frame.pack(pady=10)
        
        ttk.Label(date_frame, text="日期 (YYYY-MM-DD):").grid(row=0, column=0, padx=5)
        date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        date_entry = ttk.Entry(date_frame, textvariable=date_var, width=15)
        date_entry.grid(row=0, column=1, padx=5)
        
        # 快捷按钮
        quick_frame = ttk.Frame(date_dialog)
        quick_frame.pack(pady=10)
        
        ttk.Button(quick_frame, text="今天", 
                  command=lambda: date_var.set(datetime.now().strftime('%Y-%m-%d'))).pack(side=tk.LEFT, padx=5)
        ttk.Button(quick_frame, text="昨天", 
                  command=lambda: date_var.set((datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d'))).pack(side=tk.LEFT, padx=5)
        
        def start_scrape():
            selected_date = date_var.get()
            date_dialog.destroy()
            self._scrape_schedules_for_date(selected_date)
        
        ttk.Button(date_dialog, text="开始爬取", command=start_scrape, width=20).pack(pady=15)
    
    def _scrape_schedules_for_date(self, date):
        """爬取指定日期的排班数据"""
        def task():
            try:
                from datetime import timedelta
                self.set_status(f"正在爬取 {date} 的排班数据...")
                self.log("=" * 60)
                self.log(f"开始爬取 {date} 的排班数据", "info")
                self.log("=" * 60)
                
                # 1. 获取路线数据
                self.log("\n1️⃣ 获取路线数据...", "info")
                routes = self.real_scraper.get_all_routes(date=date, per_page=100)
                self.log(f"✓ 获取到 {len(routes)} 条路线", "success")
                
                if len(routes) == 0:
                    self.log(f"\n⚠️ {date} 没有路线数据", "warning")
                    messagebox.showwarning("提示", f"{date} 没有找到路线数据")
                    self.set_status("就绪")
                    return
                
                # 2. 获取所有司机信息
                self.log("\n2️⃣ 获取司机信息...", "info")
                driver_ids = set(r.get('driver_id') for r in routes if r.get('driver_id'))
                drivers_info = {}
                for driver_id in driver_ids:
                    try:
                        driver_data = self.api_client.get(f'/drivers/{driver_id}')
                        driver = driver_data.get('driver', {})
                        first_name = driver.get('first_name', '')
                        last_name = driver.get('last_name', '')
                        drivers_info[driver_id] = f"{first_name} {last_name}".strip()
                    except Exception as e:
                        self.log(f"  ⚠️ 获取司机 {driver_id} 信息失败: {e}", "warning")
                        drivers_info[driver_id] = f"司机{driver_id}"
                
                self.log(f"✓ 获取了 {len(drivers_info)} 位司机信息", "success")
                
                # 3. 提取司机排班
                self.log("\n3️⃣ 分析司机排班...", "info")
                from collections import defaultdict
                schedules = defaultdict(lambda: {
                    'driver_id': None,
                    'driver_name': None,
                    'date': date,
                    'routes': [],
                    'total_hours': 0,
                    'start_time': None,
                    'end_time': None
                })
                
                for route in routes:
                    driver_id = route.get('driver_id')
                    if not driver_id:
                        continue
                    
                    driver_name = drivers_info.get(driver_id, f"司机{driver_id}")
                    from_dt = route.get('from_datetime', '')
                    to_dt = route.get('to_datetime', '')
                    
                    schedules[driver_id]['driver_id'] = driver_id
                    schedules[driver_id]['driver_name'] = driver_name.strip()
                    schedules[driver_id]['routes'].append({
                        'route_id': route.get('id'),
                        'from_time': from_dt,
                        'to_time': to_dt,
                        'status': route.get('status')
                    })
                    
                    # 计算工时
                    if from_dt and to_dt:
                        try:
                            from dateutil import parser
                            start = parser.parse(from_dt)
                            end = parser.parse(to_dt)
                            hours = (end - start).total_seconds() / 3600
                            schedules[driver_id]['total_hours'] += hours
                            
                            if not schedules[driver_id]['start_time'] or start < parser.parse(schedules[driver_id]['start_time']):
                                schedules[driver_id]['start_time'] = from_dt
                            if not schedules[driver_id]['end_time'] or end > parser.parse(schedules[driver_id]['end_time']):
                                schedules[driver_id]['end_time'] = to_dt
                        except:
                            pass
                
                self.log(f"✓ 分析了 {len(schedules)} 位司机的排班", "success")
                
                # 4. 导出Excel
                self.log("\n4️⃣ 导出Excel...", "info")
                import pandas as pd
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_file = os.path.join(config.DATA_DIR, f'排班数据_{date}_{timestamp}.xlsx')
                
                # 准备排班摘要数据
                summary_data = []
                for driver_id, sched in schedules.items():
                    summary_data.append({
                        '司机ID': driver_id,
                        '司机姓名': sched['driver_name'],
                        '日期': date,
                        '路线数': len(sched['routes']),
                        '总工时': round(sched['total_hours'], 1),
                        '开始时间': sched['start_time'],
                        '结束时间': sched['end_time']
                    })
                
                # 准备详细路线数据
                detail_data = []
                for driver_id, sched in schedules.items():
                    for route in sched['routes']:
                        detail_data.append({
                            '司机ID': driver_id,
                            '司机姓名': sched['driver_name'],
                            '路线ID': route['route_id'],
                            '开始时间': route['from_time'],
                            '结束时间': route['to_time'],
                            '状态': route['status']
                        })
                
                # 保存Excel
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='排班摘要', index=False)
                    pd.DataFrame(detail_data).to_excel(writer, sheet_name='路线明细', index=False)
                    pd.DataFrame(routes).to_excel(writer, sheet_name='原始数据', index=False)
                
                self.log(f"✓ Excel已保存: {excel_file}", "success")
                
                # 保存到last_data
                self.last_data = {
                    'schedules': list(schedules.values()),
                    'routes': routes,
                    'date': date
                }
                
                self.log("\n" + "=" * 60)
                self.log("✓ 完成！", "success")
                self.log(f"日期: {date}", "info")
                self.log(f"司机数: {len(schedules)} 位", "info")
                self.log(f"路线数: {len(routes)} 条", "info")
                self.log("=" * 60)
                
                self.set_status("就绪")
                messagebox.showinfo("完成", f"排班数据爬取完成！\n\n"
                                   f"日期: {date}\n"
                                   f"司机: {len(schedules)} 位\n"
                                   f"路线: {len(routes)} 条\n\n"
                                   f"文件: {excel_file}")
                
            except Exception as e:
                import traceback
                self.log(f"✗ 爬取失败: {e}", "error")
                self.log(traceback.format_exc(), "error")
                self.set_status("就绪")
                messagebox.showerror("错误", f"爬取失败:\n{e}")
        
        threading.Thread(target=task, daemon=True).start()
    
    def generate_billing(self):
        """生成账单（finished和no_show订单，按司机分组）"""
        # 创建日期选择对话框
        date_dialog = tk.Toplevel(self.root)
        date_dialog.title("选择账单日期")
        date_dialog.geometry("350x200")
        date_dialog.transient(self.root)
        date_dialog.grab_set()
        
        # 居中显示
        date_dialog.update_idletasks()
        x = (date_dialog.winfo_screenwidth() // 2) - (350 // 2)
        y = (date_dialog.winfo_screenheight() // 2) - (200 // 2)
        date_dialog.geometry(f"350x200+{x}+{y}")
        
        ttk.Label(date_dialog, text="选择要生成账单的日期:", font=("Arial", 11)).pack(pady=15)
        
        # 日期输入框
        date_frame = ttk.Frame(date_dialog)
        date_frame.pack(pady=10)
        
        ttk.Label(date_frame, text="日期 (YYYY-MM-DD):").grid(row=0, column=0, padx=5)
        date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        date_entry = ttk.Entry(date_frame, textvariable=date_var, width=15)
        date_entry.grid(row=0, column=1, padx=5)
        
        # 快捷按钮
        quick_frame = ttk.Frame(date_dialog)
        quick_frame.pack(pady=10)
        
        ttk.Button(quick_frame, text="今天", 
                  command=lambda: date_var.set(datetime.now().strftime('%Y-%m-%d'))).pack(side=tk.LEFT, padx=5)
        ttk.Button(quick_frame, text="昨天", 
                  command=lambda: date_var.set((datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d'))).pack(side=tk.LEFT, padx=5)
        
        def start_generate():
            selected_date = date_var.get()
            date_dialog.destroy()
            self._generate_billing_for_date(selected_date)
        
        ttk.Button(date_dialog, text="生成账单", command=start_generate, width=20).pack(pady=15)
    
    def _generate_billing_for_date(self, date):
        """生成指定日期的账单"""
        def task():
            try:
                self.set_status(f"正在生成 {date} 的账单...")
                self.log("=" * 60)
                self.log(f"开始生成 {date} 的账单", "info")
                self.log("=" * 60)
                
                # 获取finished和no_show状态的订单
                self.log("\n1️⃣ 获取已完成和未到达订单...", "info")
                rides = self.real_scraper.get_all_rides(
                    date=date, 
                    per_page=500, 
                    statuses='finished,no_show'
                )
                self.log(f"✓ 获取到 {len(rides)} 条订单", "success")
                
                if len(rides) == 0:
                    self.log(f"\n⚠️ {date} 没有符合条件的订单", "warning")
                    messagebox.showwarning("提示", f"{date} 没有找到finished或no_show状态的订单")
                    self.set_status("就绪")
                    return
                
                # 获取订单详细信息（包含价格）
                self.log("\n2️⃣ 获取订单详细信息（价格）...", "info")
                detailed_rides = []
                for idx, ride in enumerate(rides, 1):
                    try:
                        ride_id = ride.get('id')
                        detail = self.api_client.get(f'/rides/{ride_id}')
                        ride_detail = detail.get('ride', {})
                        
                        # 合并基本信息和详细信息
                        ride['driver_net'] = ride_detail.get('driver_net', 0)
                        ride['vendor_amount'] = ride_detail.get('vendor_amount', 0)
                        ride['distance'] = ride_detail.get('distance', 0)
                        ride['duration'] = ride_detail.get('duration', 0)
                        detailed_rides.append(ride)
                        
                        if idx % 50 == 0:
                            self.log(f"  已处理 {idx}/{len(rides)} 条订单...", "info")
                    except Exception as e:
                        self.log(f"  ⚠️ 获取订单 {ride.get('id')} 详情失败: {e}", "warning")
                        detailed_rides.append(ride)
                
                self.log(f"✓ 已获取 {len(detailed_rides)} 条订单详情", "success")
                
                # 按司机分组
                self.log("\n3️⃣ 按司机分组订单...", "info")
                from collections import defaultdict
                driver_orders = defaultdict(list)
                
                for ride in detailed_rides:
                    driver_id = ride.get('driver_id')
                    if driver_id:
                        driver_orders[driver_id].append(ride)
                
                self.log(f"✓ 共有 {len(driver_orders)} 位司机", "success")
                
                # 导出Excel
                self.log("\n4️⃣ 导出账单Excel...", "info")
                import pandas as pd
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_file = os.path.join(config.DATA_DIR, f'账单_{date}_{timestamp}.xlsx')
                
                # 准备所有司机的订单数据
                all_rows = []
                
                # 按司机名字排序
                driver_list = []
                for driver_id, orders in driver_orders.items():
                    first_ride = orders[0]
                    driver_first = first_ride.get('driver_first_name', '')
                    driver_last = first_ride.get('driver_last_name', '')
                    driver_name = f"{driver_first} {driver_last}".strip()
                    driver_list.append((driver_name, driver_id, orders))
                
                driver_list.sort(key=lambda x: x[0])  # 按名字排序
                
                for driver_name, driver_id, orders in driver_list:
                    # 计算该司机的总收入
                    total_earnings = sum(float(order.get('driver_net', 0) or 0) for order in orders)
                    
                    # 司机标题行
                    all_rows.append({
                        '司机姓名': driver_name,
                        '订单数': len(orders),
                        '总收入': f"${total_earnings:.2f}",
                        '订单ID': '',
                        '接客时间': '',
                        '接客地点': '',
                        '送达地点': '',
                        '乘客姓名': '',
                        '价格': '',
                        '里程': '',
                        '状态': ''
                    })
                    
                    # 该司机的所有订单
                    for ride in orders:
                        driver_net = float(ride.get('driver_net', 0) or 0)
                        distance = float(ride.get('distance', 0) or 0)
                        all_rows.append({
                            '司机姓名': driver_name,  # 每行都显示司机名字
                            '订单数': '',
                            '总收入': '',
                            '订单ID': ride.get('id', ''),
                            '接客时间': ride.get('pickup_at', ''),
                            '接客地点': ride.get('start_address', ''),
                            '送达地点': ride.get('destination_address', ''),
                            '乘客姓名': f"{ride.get('first_name', '')} {ride.get('last_name', '')}".strip(),
                            '价格': f"${driver_net:.2f}" if driver_net > 0 else '',
                            '里程': f"{distance:.1f} mi" if distance > 0 else '',
                            '状态': ride.get('status', '')
                        })
                    
                    # 空行分隔
                    all_rows.append({
                        '司机姓名': '',
                        '订单数': '',
                        '总收入': '',
                        '订单ID': '',
                        '接客时间': '',
                        '接客地点': '',
                        '送达地点': '',
                        '乘客姓名': '',
                        '价格': '',
                        '里程': '',
                        '状态': ''
                    })
                
                # 保存到Excel
                df = pd.DataFrame(all_rows)
                df.to_excel(excel_file, index=False, sheet_name='账单')
                
                # 美化Excel格式
                self.log("\n5️⃣ 格式化Excel...", "info")
                wb = load_workbook(excel_file)
                ws = wb.active
                
                # 设置列宽
                ws.column_dimensions['A'].width = 20  # 司机姓名
                ws.column_dimensions['B'].width = 10  # 订单数
                ws.column_dimensions['C'].width = 12  # 总收入
                ws.column_dimensions['D'].width = 12  # 订单ID
                ws.column_dimensions['E'].width = 20  # 接客时间
                ws.column_dimensions['F'].width = 45  # 接客地点
                ws.column_dimensions['G'].width = 45  # 送达地点
                ws.column_dimensions['H'].width = 20  # 乘客姓名
                ws.column_dimensions['I'].width = 12  # 价格
                ws.column_dimensions['J'].width = 12  # 里程
                ws.column_dimensions['K'].width = 12  # 状态
                
                # 样式定义
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                driver_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                driver_font = Font(bold=True, size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                center_align = Alignment(horizontal='center', vertical='center')
                
                # 格式化表头
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border
                
                # 格式化数据行
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    # 检查是否是司机标题行（有司机姓名且有订单数但没有订单ID）
                    if row[0].value and row[1].value and not row[3].value:  # 司机姓名有值，订单数有值，订单ID无值（订单ID在第4列）
                        for cell in row:
                            cell.fill = driver_fill
                            cell.font = driver_font
                            cell.alignment = center_align
                            cell.border = border
                    else:
                        for cell in row:
                            cell.alignment = Alignment(vertical='center')
                            cell.border = border
                
                wb.save(excel_file)
                self.log(f"✓ Excel已保存: {excel_file}", "success")
                
                # 统计信息
                self.log("\n" + "=" * 60)
                self.log("✓ 完成！", "success")
                self.log(f"日期: {date}", "info")
                self.log(f"订单总数: {len(rides)} 条", "info")
                self.log(f"司机数: {len(driver_orders)} 位", "info")
                for status in ['finished', 'no_show']:
                    count = sum(1 for r in rides if r.get('status') == status)
                    if count > 0:
                        self.log(f"  {status}: {count} 条", "info")
                self.log("=" * 60)
                
                self.set_status("就绪")
                messagebox.showinfo("完成", f"账单生成完成！\n\n"
                                   f"日期: {date}\n"
                                   f"订单: {len(rides)} 条\n"
                                   f"司机: {len(driver_orders)} 位\n\n"
                                   f"文件: {excel_file}")
                
            except Exception as e:
                import traceback
                self.log(f"✗ 生成账单失败: {e}", "error")
                self.log(traceback.format_exc(), "error")
                self.set_status("就绪")
                messagebox.showerror("错误", f"生成账单失败:\n{e}")
        
        threading.Thread(target=task, daemon=True).start()
    
    def scrape_orders_only(self):
        """只爬取订单数据"""
        # 创建日期选择对话框
        date_dialog = tk.Toplevel(self.root)
        date_dialog.title("选择爬取日期")
        date_dialog.geometry("350x200")
        date_dialog.transient(self.root)
        date_dialog.grab_set()
        
        # 居中显示
        date_dialog.update_idletasks()
        x = (date_dialog.winfo_screenwidth() // 2) - (350 // 2)
        y = (date_dialog.winfo_screenheight() // 2) - (200 // 2)
        date_dialog.geometry(f"350x200+{x}+{y}")
        
        ttk.Label(date_dialog, text="选择要爬取订单数据的日期:", font=("Arial", 11)).pack(pady=15)
        
        # 日期输入框
        date_frame = ttk.Frame(date_dialog)
        date_frame.pack(pady=10)
        
        ttk.Label(date_frame, text="日期 (YYYY-MM-DD):").grid(row=0, column=0, padx=5)
        date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        date_entry = ttk.Entry(date_frame, textvariable=date_var, width=15)
        date_entry.grid(row=0, column=1, padx=5)
        
        # 快捷按钮
        quick_frame = ttk.Frame(date_dialog)
        quick_frame.pack(pady=10)
        
        ttk.Button(quick_frame, text="今天", 
                  command=lambda: date_var.set(datetime.now().strftime('%Y-%m-%d'))).pack(side=tk.LEFT, padx=5)
        ttk.Button(quick_frame, text="昨天", 
                  command=lambda: date_var.set((datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d'))).pack(side=tk.LEFT, padx=5)
        
        def start_scrape():
            selected_date = date_var.get()
            date_dialog.destroy()
            self._scrape_orders_for_date(selected_date)
        
        ttk.Button(date_dialog, text="开始爬取", command=start_scrape, width=20).pack(pady=15)
    
    def _scrape_orders_for_date(self, date):
        """爬取指定日期的订单数据"""
        def task():
            try:
                self.set_status(f"正在爬取 {date} 的订单数据...")
                self.log("=" * 60)
                self.log(f"开始爬取 {date} 的订单数据", "info")
                self.log("=" * 60)
                
                # 获取订单数据（使用rides API）
                self.log("\n1️⃣ 获取订单数据...", "info")
                rides = self.real_scraper.get_all_rides(date=date, per_page=500, statuses='')
                self.log(f"✓ 获取到 {len(rides)} 条订单", "success")
                
                if len(rides) == 0:
                    self.log(f"\n⚠️ {date} 没有订单数据", "warning")
                    messagebox.showwarning("提示", f"{date} 没有找到订单数据")
                    self.set_status("就绪")
                    return
                
                # 获取订单详细信息（包含价格）
                self.log("\n2️⃣ 获取订单详细信息（价格）...", "info")
                detailed_rides = []
                for idx, ride in enumerate(rides, 1):
                    try:
                        ride_id = ride.get('id')
                        detail = self.api_client.get(f'/rides/{ride_id}')
                        ride_detail = detail.get('ride', {})
                        
                        # 合并基本信息和详细信息
                        ride['driver_net'] = ride_detail.get('driver_net', 0)
                        ride['distance'] = ride_detail.get('distance', 0)
                        ride['duration'] = ride_detail.get('duration', 0)
                        detailed_rides.append(ride)
                        
                        if idx % 100 == 0:
                            self.log(f"  已处理 {idx}/{len(rides)} 条订单...", "info")
                    except Exception as e:
                        self.log(f"  ⚠️ 获取订单 {ride.get('id')} 详情失败: {e}", "warning")
                        detailed_rides.append(ride)
                
                self.log(f"✓ 已获取 {len(detailed_rides)} 条订单详情", "success")
                
                # 统计订单信息
                self.log("\n3️⃣ 统计订单信息...", "info")
                from collections import Counter
                
                status_count = Counter(r.get('status') for r in detailed_rides)
                driver_count = len(set(r.get('driver_id') for r in detailed_rides if r.get('driver_id')))
                
                self.log(f"  订单总数: {len(detailed_rides)}", "info")
                self.log(f"  司机数: {driver_count}", "info")
                for status, count in status_count.items():
                    self.log(f"  状态 '{status}': {count} 条", "info")
                
                # 导出Excel
                self.log("\n4️⃣ 导出Excel...", "info")
                import pandas as pd
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_file = os.path.join(config.DATA_DIR, f'订单数据_{date}_{timestamp}.xlsx')
                
                # 整理订单数据
                order_data = []
                for ride in detailed_rides:
                    driver_first = ride.get('driver_first_name', '')
                    driver_last = ride.get('driver_last_name', '')
                    driver_name = f"{driver_first} {driver_last}".strip()
                    
                    passenger_first = ride.get('first_name', '')
                    passenger_last = ride.get('last_name', '')
                    passenger_name = f"{passenger_first} {passenger_last}".strip()
                    
                    driver_net = float(ride.get('driver_net', 0) or 0)
                    distance = float(ride.get('distance', 0) or 0)
                    
                    order_data.append({
                        '订单ID': ride.get('id'),
                        '日期': date,
                        '司机ID': ride.get('driver_id'),
                        '司机姓名': driver_name,
                        '接客时间': ride.get('pickup_at', ''),
                        '接客地址': ride.get('start_address', ''),
                        '送达地址': ride.get('destination_address', ''),
                        '乘客姓名': passenger_name,
                        '价格': f"${driver_net:.2f}" if driver_net > 0 else '',
                        '里程': f"{distance:.1f} mi" if distance > 0 else '',
                        '状态': ride.get('status'),
                    })
                
                # 保存Excel
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    pd.DataFrame(order_data).to_excel(writer, sheet_name='订单列表', index=False)
                    
                    # 添加统计摘要
                    total_earnings = sum(float(r.get('driver_net', 0) or 0) for r in detailed_rides)
                    summary_data = [{
                        '日期': date,
                        '订单总数': len(detailed_rides),
                        '司机数': driver_count,
                        '总收入': f"${total_earnings:.2f}",
                        **{f'状态_{k}': v for k, v in status_count.items()}
                    }]
                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='统计摘要', index=False)
                    
                    # 原始数据
                    pd.DataFrame(detailed_rides).to_excel(writer, sheet_name='原始数据', index=False)
                
                self.log(f"✓ Excel已保存: {excel_file}", "success")
                
                # 保存到last_data
                self.last_data = {
                    'orders': detailed_rides,
                    'date': date
                }
                
                total_earnings = sum(float(r.get('driver_net', 0) or 0) for r in detailed_rides)
                
                self.log("\n" + "=" * 60)
                self.log("✓ 完成！", "success")
                self.log(f"日期: {date}", "info")
                self.log(f"订单数: {len(detailed_rides)} 条", "info")
                self.log(f"司机数: {driver_count} 位", "info")
                self.log(f"总收入: ${total_earnings:.2f}", "info")
                self.log("=" * 60)
                
                self.set_status("就绪")
                messagebox.showinfo("完成", f"订单数据爬取完成！\n\n"
                                   f"日期: {date}\n"
                                   f"订单: {len(detailed_rides)} 条\n"
                                   f"司机: {driver_count} 位\n"
                                   f"总收入: ${total_earnings:.2f}\n\n"
                                   f"文件: {excel_file}")
                
            except Exception as e:
                import traceback
                self.log(f"✗ 爬取失败: {e}", "error")
                self.log(traceback.format_exc(), "error")
                self.set_status("就绪")
                messagebox.showerror("错误", f"爬取失败:\n{e}")
        
        threading.Thread(target=task, daemon=True).start()
    
    def quick_test_scrape(self):
        """快速测试（10条数据）"""
        def task():
            try:
                self.set_status("快速测试中...")
                self.log("=" * 60)
                self.log("快速测试 - 10位司机", "info")
                self.log("=" * 60)
                
                drivers = self.real_scraper.get_all_drivers(per_page=10)
                self.log(f"✓ 获取到 {len(drivers[:10])} 位司机", "success")
                
                for i, driver in enumerate(drivers[:10], 1):
                    self.log(f"{i}. {driver.get('first_name', '')} {driver.get('last_name', '')} (ID: {driver.get('id')})")
                
                self.set_status("就绪")
                messagebox.showinfo("测试完成", f"成功获取 {len(drivers[:10])} 位司机数据")
                
            except Exception as e:
                self.log(f"✗ 测试失败: {e}", "error")
                self.set_status("就绪")
        
        threading.Thread(target=task, daemon=True).start()
    
    def view_schedules(self):
        """查看工作排班"""
        def view():
            self.set_status("显示排班信息...")
            try:
                if not self.last_data or 'schedules' not in self.last_data:
                    self.log("请先爬取数据", "warning")
                    messagebox.showinfo("提示", "请先使用'爬取完整数据'功能获取排班信息")
                    return
                
                schedules = self.last_data['schedules']
                self.log("\n" + "=" * 60)
                self.log(f"工作排班统计 (共 {len(schedules)} 位司机)", "info")
                self.log("=" * 60)
                
                # 按路线数排序
                sorted_schedules = sorted(schedules, key=lambda x: x.get('total_routes', 0), reverse=True)
                
                for i, schedule in enumerate(sorted_schedules[:15], 1):
                    self.log(f"\n{i}. {schedule.get('driver_name', '未知')}")
                    self.log(f"   电话: {schedule.get('phone')} | 车辆: {schedule.get('plate')}")
                    self.log(f"   路线: {schedule.get('total_routes')} 条 | 工时: {schedule.get('total_hours', 0):.1f} 小时")
                    self.log(f"   时段: {schedule.get('earliest_start')} ~ {schedule.get('latest_end')}")
                
                if len(schedules) > 15:
                    self.log(f"\n... 还有 {len(schedules) - 15} 位司机")
                
            except Exception as e:
                self.log(f"✗ 显示失败: {e}", "error")
            finally:
                self.set_status("就绪")
        
        threading.Thread(target=view, daemon=True).start()
    
    def show_statistics(self):
        """显示数据统计"""
        def show():
            try:
                if not self.last_data:
                    messagebox.showinfo("提示", "请先爬取数据")
                    return
                
                self.log("\n" + "=" * 60)
                self.log("📊 数据统计", "info")
                self.log("=" * 60)
                
                metadata = self.last_data.get('metadata', {})
                drivers = self.last_data.get('drivers', [])
                routes = self.last_data.get('routes', [])
                schedules = self.last_data.get('schedules', [])
                
                self.log(f"\n数据概览:")
                self.log(f"  司机总数: {len(drivers)} 位")
                self.log(f"  路线总数: {len(routes)} 条")
                self.log(f"  有排班司机: {len(schedules)} 位")
                
                if schedules:
                    total_routes = sum(s.get('total_routes', 0) for s in schedules)
                    total_hours = sum(s.get('total_hours', 0) for s in schedules)
                    self.log(f"\n工作统计:")
                    self.log(f"  总路线: {total_routes} 条")
                    self.log(f"  总工时: {total_hours:.1f} 小时")
                    self.log(f"  平均每人: {total_routes/len(schedules):.1f} 条路线")
                
                if routes:
                    from collections import Counter
                    statuses = Counter(r.get('status') for r in routes)
                    self.log(f"\n路线状态:")
                    for status, count in statuses.most_common():
                        self.log(f"  {status}: {count} 条")
                
                self.log("\n" + "=" * 60)
                
            except Exception as e:
                self.log(f"✗ 统计失败: {e}", "error")
        
        threading.Thread(target=show, daemon=True).start()
    
    def export_json(self):
        """导出为JSON"""
        try:
            if not self.last_data:
                # 尝试加载保存的数据
                self.last_data = self.scraper.load_data()
            
            if not self.last_data:
                messagebox.showwarning("提示", "请先爬取数据")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                initialfile=f"rpa_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            )
            
            if filename:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(self.last_data, f, ensure_ascii=False, indent=2)
                
                self.log(f"✓ 数据已导出到: {filename}", "success")
                messagebox.showinfo("成功", f"数据已导出到:\n{filename}")
        
        except Exception as e:
            self.log(f"✗ 导出失败: {e}", "error")
            messagebox.showerror("错误", f"导出失败: {e}")
    
    def scrape_complete_data(self):
        """爬取完整数据（使用真实API）"""
        def task():
            try:
                self.set_status("正在爬取完整数据...")
                self.log("=" * 60)
                self.log("开始爬取完整数据（真实API）", "info")
                self.log("=" * 60)
                
                # 询问模式
                response = messagebox.askyesnocancel(
                    "选择爬取模式",
                    "选择数据爬取模式:\n\n"
                    "是 - 快速测试 (前50条司机 + 今天路线)\n"
                    "否 - 完整爬取 (所有司机 + 今天路线)\n"
                    "取消 - 返回"
                )
                
                if response is None:
                    self.set_status("就绪")
                    return
                
                per_page = 50 if response else 100
                mode = "测试模式" if response else "完整模式"
                self.log(f"📊 {mode}", "info")
                
                result = {
                    'timestamp': datetime.now().isoformat(),
                    'drivers': [],
                    'routes': [],
                    'metadata': {}
                }
                
                # 1. 爬取司机数据
                self.log("\n" + "-" * 60)
                self.log("正在爬取司机数据...", "info")
                
                def driver_progress(current, total, name):
                    self.log(f"  [{current}/{total}] {name}")
                
                result['drivers'] = self.real_scraper.get_all_drivers(
                    per_page=per_page,
                    progress_callback=driver_progress
                )
                result['metadata']['total_drivers'] = len(result['drivers'])
                
                self.log(f"✓ 司机数据: {len(result['drivers'])} 位", "success")
                
                # 2. 爬取路线数据
                self.log("\n" + "-" * 60)
                self.log("正在爬取路线数据...", "info")
                
                today = datetime.now().strftime('%Y-%m-%d')
                
                def route_progress(current, total, name):
                    self.log(f"  [{current}/{total}] {name}")
                
                result['routes'] = self.real_scraper.get_all_routes(
                    date=today,
                    per_page=per_page,
                    progress_callback=route_progress
                )
                result['metadata']['total_routes'] = len(result['routes'])
                result['metadata']['route_date'] = today
                
                self.log(f"✓ 路线数据: {len(result['routes'])} 条", "success")
                
                # 3. 保存数据
                self.last_data = result
                
                self.log("\n" + "-" * 60)
                self.log("正在保存数据...", "info")
                
                json_file = self.real_scraper.save_to_json(result)
                self.log(f"✓ JSON: {json_file}", "success")
                
                excel_file = self.real_scraper.export_to_excel(result)
                self.log(f"✓ Excel: {excel_file}", "success")
                
                # 4. 显示摘要
                self.log("\n" + "=" * 60)
                self.log("✓ 爬取完成！", "success")
                self.log("=" * 60)
                self.log(f"司机总数: {result['metadata']['total_drivers']}", "info")
                self.log(f"路线总数: {result['metadata']['total_routes']}", "info")
                self.log(f"路线日期: {result['metadata']['route_date']}", "info")
                self.log("=" * 60)
                
                # 显示样例
                if result['drivers']:
                    self.log("\n司机数据样例 (前3位):", "info")
                    for i, d in enumerate(result['drivers'][:3], 1):
                        self.log(f"  {i}. {d.get('first_name')} {d.get('last_name')}")
                        self.log(f"     电话: {d.get('phone_number')} | 车辆: {d.get('plate_number')}")
                
                self.set_status("就绪")
                
                messagebox.showinfo(
                    "爬取完成",
                    f"数据爬取完成！\n\n"
                    f"司机: {result['metadata']['total_drivers']} 位\n"
                    f"路线: {result['metadata']['total_routes']} 条\n"
                    f"日期: {result['metadata']['route_date']}\n\n"
                    f"文件已保存:\n"
                    f"JSON: {json_file}\n"
                    f"Excel: {excel_file}"
                )
                
            except Exception as e:
                import traceback
                error_msg = traceback.format_exc()
                self.log(f"✗ 爬取失败: {e}", "error")
                self.log(error_msg, "error")
                self.set_status("就绪")
                messagebox.showerror("错误", f"爬取失败:\n{e}")
        
        threading.Thread(target=task, daemon=True).start()
    
    def export_excel(self):
        """导出为Excel"""
        try:
            if not self.last_data:
                self.last_data = self.scraper.load_data()
            
            if not self.last_data:
                messagebox.showwarning("提示", "请先爬取数据")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"rpa_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if filename:
                import pandas as pd
                
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # 导出司机数据
                    if self.last_data.get('drivers'):
                        df_drivers = pd.DataFrame(self.last_data['drivers'])
                        df_drivers.to_excel(writer, sheet_name='司机', index=False)
                    
                    # 导出车辆数据
                    if self.last_data.get('vehicles'):
                        df_vehicles = pd.DataFrame(self.last_data['vehicles'])
                        df_vehicles.to_excel(writer, sheet_name='车辆', index=False)
                    
                    # 导出排班数据
                    if self.last_data.get('schedules'):
                        df_schedules = pd.DataFrame(self.last_data['schedules'])
                        df_schedules.to_excel(writer, sheet_name='排班', index=False)
                
                self.log(f"✓ 数据已导出到: {filename}", "success")
                messagebox.showinfo("成功", f"数据已导出到:\n{filename}")
        
        except ImportError:
            messagebox.showerror("错误", "缺少pandas或openpyxl库\n请运行: pip install pandas openpyxl")
        except Exception as e:
            self.log(f"✗ 导出失败: {e}", "error")
            messagebox.showerror("错误", f"导出失败: {e}")
    
    def open_data_folder(self):
        """打开数据目录"""
        try:
            data_dir = os.path.abspath(config.DATA_DIR)
            if not os.path.exists(data_dir):
                os.makedirs(data_dir)
            
            os.startfile(data_dir)
            self.log(f"✓ 已打开数据目录: {data_dir}", "info")
        except Exception as e:
            self.log(f"✗ 打开失败: {e}", "error")
    
    def clean_data_folder(self):
        """清理数据目录"""
        try:
            data_dir = os.path.abspath(config.DATA_DIR)
            if not os.path.exists(data_dir):
                messagebox.showinfo("提示", "数据目录不存在或为空")
                return
            
            # 统计文件
            files = [f for f in os.listdir(data_dir) if os.path.isfile(os.path.join(data_dir, f))]
            if not files:
                messagebox.showinfo("提示", "数据目录已经是空的")
                return
            
            # 显示确认对话框
            file_list = "\n".join(files[:10])  # 最多显示前10个文件
            if len(files) > 10:
                file_list += f"\n... 还有 {len(files) - 10} 个文件"
            
            result = messagebox.askyesno(
                "确认清理",
                f"确定要删除以下 {len(files)} 个文件吗？\n\n{file_list}\n\n此操作不可恢复！",
                icon='warning'
            )
            
            if not result:
                self.log("✓ 已取消清理操作", "info")
                return
            
            # 执行删除
            self.log("=" * 60)
            self.log("开始清理数据目录...", "info")
            deleted_count = 0
            failed_count = 0
            
            for filename in files:
                file_path = os.path.join(data_dir, filename)
                try:
                    os.remove(file_path)
                    deleted_count += 1
                    self.log(f"✓ 已删除: {filename}", "success")
                except Exception as e:
                    failed_count += 1
                    self.log(f"✗ 删除失败: {filename} - {e}", "error")
            
            self.log("=" * 60)
            self.log(f"✓ 清理完成！成功删除 {deleted_count} 个文件", "success")
            if failed_count > 0:
                self.log(f"⚠️ 有 {failed_count} 个文件删除失败", "warning")
            
            messagebox.showinfo("完成", f"清理完成！\n\n成功删除: {deleted_count} 个文件\n失败: {failed_count} 个文件")
            
        except Exception as e:
            self.log(f"✗ 清理失败: {e}", "error")
            messagebox.showerror("错误", f"清理失败: {e}")
    
    def show_dispatch_dialog(self):
        """显示派工对话框 (Assign Driver)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("派工 (Assign)")
        dialog.geometry("450x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (200 // 2)
        dialog.geometry(f"450x200+{x}+{y}")
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        # 输入字段
        ttk.Label(frame, text="订单ID (Ride ID):", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W, pady=10)
        ride_id_entry = ttk.Entry(frame, width=30)
        ride_id_entry.grid(row=0, column=1, pady=10, padx=10)
        
        ttk.Label(frame, text="司机ID (Driver ID):", font=("Arial", 10)).grid(row=1, column=0, sticky=tk.W, pady=10)
        driver_id_entry = ttk.Entry(frame, width=30)
        driver_id_entry.grid(row=1, column=1, pady=10, padx=10)
        
        def submit():
            try:
                ride_id = int(ride_id_entry.get().strip())
                driver_id = int(driver_id_entry.get().strip())
                
                self.log("=" * 60)
                self.log(f"开始派工: 订单 {ride_id} -> 司机 {driver_id}", "info")
                
                result = self.dispatcher.assign_driver(ride_id, driver_id)
                
                self.log(f"✓ 派工成功", "success")
                self.log(f"响应: {result}")
                self.log("=" * 60)
                
                messagebox.showinfo("成功", f"派工成功！\n\n订单ID: {ride_id}\n司机ID: {driver_id}")
                dialog.destroy()
                
            except ValueError:
                messagebox.showerror("错误", "请输入有效的数字ID")
            except Exception as e:
                self.log(f"✗ 派工失败: {e}", "error")
                messagebox.showerror("错误", f"派工失败:\n{e}")
        
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=20)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="确认派工", command=submit).pack(side=tk.LEFT, padx=5)
    
    def show_withdraw_dialog(self):
        """显示退工对话框 (Revive - Cancel Ride) - 按司机ID和时间段"""
        dialog = tk.Toplevel(self.root)
        dialog.title("退工 (Revive)")
        dialog.geometry("500x250")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (250 // 2)
        dialog.geometry(f"500x250+{x}+{y}")
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="司机ID (Driver ID):", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W, pady=8)
        driver_id_entry = ttk.Entry(frame, width=30)
        driver_id_entry.grid(row=0, column=1, pady=8, padx=10)
        
        ttk.Label(frame, text="日期 (Date):", font=("Arial", 10)).grid(row=1, column=0, sticky=tk.W, pady=8)
        date_entry = ttk.Entry(frame, width=30)
        date_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        date_entry.grid(row=1, column=1, pady=8, padx=10)
        
        ttk.Label(frame, text="时间段 (Time Range):", font=("Arial", 10)).grid(row=2, column=0, sticky=tk.W, pady=8)
        time_entry = ttk.Entry(frame, width=30)
        time_entry.insert(0, "00:00-23:59")
        time_entry.grid(row=2, column=1, pady=8, padx=10)
        
        ttk.Label(frame, text="取消原因:", font=("Arial", 10)).grid(row=3, column=0, sticky=tk.W, pady=8)
        reason_label = ttk.Label(frame, text="driver cancel (固定)", foreground="gray")
        reason_label.grid(row=3, column=1, pady=8, padx=10, sticky=tk.W)
        
        def submit():
            try:
                driver_id = int(driver_id_entry.get().strip())
                date = date_entry.get().strip()
                time_range = time_entry.get().strip()
                reason = "driver cancel"
                
                self.log("=" * 60)
                self.log(f"开始批量退工", "info")
                self.log(f"司机ID: {driver_id}", "info")
                self.log(f"日期: {date}", "info")
                self.log(f"时间段: {time_range}", "info")
                self.log(f"原因: {reason}", "info")
                
                # 获取该司机在指定时间段的所有订单
                from_time, to_time = time_range.split('-')
                # 使用空格匹配API返回格式: "2025-11-22 08:00:00"
                from_datetime = f"{date} {from_time.strip()}:00"
                to_datetime = f"{date} {to_time.strip()}:00"
                
                self.log(f"获取时间段: {from_datetime} ~ {to_datetime}", "info")
                
                rides = self.real_scraper.get_all_rides(
                    date=date,
                    per_page=500,
                    statuses=''
                )
                
                # 筛选该司机在指定时间段的订单
                driver_rides = []
                for r in rides:
                    if r.get('driver_id') == driver_id:
                        pickup_time = r.get('pickup_at', '')
                        # 如果有pickup_at字段，检查是否在时间范围内
                        if pickup_time:
                            # 精确比较到分钟 (format: 2025-11-22 09:00:00)
                            if from_datetime[:16] <= pickup_time[:16] <= to_datetime[:16]:
                                driver_rides.append(r)
                        else:
                            driver_rides.append(r)
                
                self.log(f"找到 {len(driver_rides)} 条该司机在指定时间段的订单", "info")
                
                if len(driver_rides) == 0:
                    messagebox.showwarning("提示", f"未找到司机 {driver_id} 在该时间段的订单")
                    return
                
                # 逐个退工
                success_count = 0
                fail_count = 0
                
                for ride in driver_rides:
                    ride_id = ride.get('id')
                    status = ride.get('status', '')
                    driver_name = f"{ride.get('driver_first_name', '')} {ride.get('driver_last_name', '')}"
                    pickup_at = ride.get('pickup_at', '')
                    
                    self.log(f"  订单 {ride_id} ({pickup_at}, 状态: {status})", "info")
                    
                    try:
                        self.dispatcher.cancel_ride(ride_id, reason)
                        success_count += 1
                        self.log(f"    ✓ 退工成功", "success")
                    except Exception as e:
                        fail_count += 1
                        error_msg = str(e)
                        if "404" in error_msg:
                            self.log(f"    ✗ 失败: 订单不允许退工 (404)", "error")
                        elif "403" in error_msg:
                            self.log(f"    ✗ 失败: 无权限 (403)", "error")
                        else:
                            self.log(f"    ✗ 失败: {e}", "error")
                
                self.log("=" * 60)
                self.log(f"✓ 批量退工完成", "success")
                self.log(f"成功: {success_count} 条, 失败: {fail_count} 条", "info")
                self.log("=" * 60)
                
                msg = f"批量退工完成！\n\n成功: {success_count} 条\n失败: {fail_count} 条"
                
                if success_count == 0 and fail_count > 0:
                    messagebox.showwarning("完成", msg + "\n\n⚠️ 所有订单退工失败\n可能原因：订单状态不允许退工")
                else:
                    messagebox.showinfo("完成", msg)
                dialog.destroy()
                
            except ValueError:
                messagebox.showerror("错误", "请输入有效的司机ID")
            except Exception as e:
                self.log(f"✗ 批量退工失败: {e}", "error")
                messagebox.showerror("错误", f"批量退工失败:\n{e}")
        
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=15)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="确认批量退工", command=submit).pack(side=tk.LEFT, padx=5)
    
    def show_transfer_dialog(self):
        """显示转派对话框 (Switch Driver) - 按司机ID和时间段"""
        dialog = tk.Toplevel(self.root)
        dialog.title("转派 (Switch)")
        dialog.geometry("500x280")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (280 // 2)
        dialog.geometry(f"500x280+{x}+{y}")
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="原司机ID (From Driver):", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W, pady=8)
        from_driver_entry = ttk.Entry(frame, width=30)
        from_driver_entry.grid(row=0, column=1, pady=8, padx=10)
        
        ttk.Label(frame, text="新司机ID (To Driver):", font=("Arial", 10)).grid(row=1, column=0, sticky=tk.W, pady=8)
        to_driver_entry = ttk.Entry(frame, width=30)
        to_driver_entry.grid(row=1, column=1, pady=8, padx=10)
        
        ttk.Label(frame, text="日期 (Date):", font=("Arial", 10)).grid(row=2, column=0, sticky=tk.W, pady=8)
        date_entry = ttk.Entry(frame, width=30)
        date_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        date_entry.grid(row=2, column=1, pady=8, padx=10)
        
        ttk.Label(frame, text="时间段 (Time Range):", font=("Arial", 10)).grid(row=3, column=0, sticky=tk.W, pady=8)
        time_entry = ttk.Entry(frame, width=30)
        time_entry.insert(0, "00:00-23:59")
        time_entry.grid(row=3, column=1, pady=8, padx=10)
        
        def submit():
            try:
                from_driver_id = int(from_driver_entry.get().strip())
                to_driver_id = int(to_driver_entry.get().strip())
                date = date_entry.get().strip()
                time_range = time_entry.get().strip()
                
                self.log("=" * 60)
                self.log(f"开始批量转派", "info")
                self.log(f"原司机ID: {from_driver_id}", "info")
                self.log(f"新司机ID: {to_driver_id}", "info")
                self.log(f"日期: {date}", "info")
                self.log(f"时间段: {time_range}", "info")
                
                # 获取该司机在指定时间段的所有订单
                from_time, to_time = time_range.split('-')
                # 使用空格匹配API返回格式: "2025-11-22 08:00:00"
                from_datetime = f"{date} {from_time.strip()}:00"
                to_datetime = f"{date} {to_time.strip()}:00"
                
                self.log(f"获取时间段: {from_datetime} ~ {to_datetime}", "info")
                
                rides = self.real_scraper.get_all_rides(
                    date=date,
                    per_page=500,
                    statuses=''
                )
                
                # 筛选该司机在指定时间段的订单
                driver_rides = []
                for r in rides:
                    if r.get('driver_id') == from_driver_id:
                        pickup_time = r.get('pickup_at', '')
                        # 如果有pickup_at字段，检查是否在时间范围内
                        if pickup_time:
                            # 精确比较到分钟 (format: 2025-11-22 09:00:00)
                            if from_datetime[:16] <= pickup_time[:16] <= to_datetime[:16]:
                                driver_rides.append(r)
                        else:
                            driver_rides.append(r)
                
                self.log(f"找到 {len(driver_rides)} 条该司机在指定时间段的订单", "info")
                
                if len(driver_rides) == 0:
                    messagebox.showwarning("提示", f"未找到司机 {from_driver_id} 在该时间段的订单")
                    return
                
                # 逐个转派
                success_count = 0
                fail_count = 0
                for ride in driver_rides:
                    try:
                        ride_id = ride.get('id')
                        self.dispatcher.transfer_driver(ride_id, to_driver_id)
                        success_count += 1
                        self.log(f"  ✓ 订单 {ride_id} 转派成功", "success")
                    except Exception as e:
                        fail_count += 1
                        self.log(f"  ✗ 订单 {ride.get('id')} 转派失败: {e}", "error")
                
                self.log("=" * 60)
                self.log(f"✓ 批量转派完成", "success")
                self.log(f"成功: {success_count} 条, 失败: {fail_count} 条", "info")
                self.log("=" * 60)
                
                messagebox.showinfo("完成", f"批量转派完成！\n\n成功: {success_count} 条\n失败: {fail_count} 条")
                dialog.destroy()
                
            except ValueError:
                messagebox.showerror("错误", "请输入有效的数字ID")
            except Exception as e:
                self.log(f"✗ 批量转派失败: {e}", "error")
                messagebox.showerror("错误", f"批量转派失败:\n{e}")
        
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=15)
        ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="确认批量转派", command=submit).pack(side=tk.LEFT, padx=5)
    
    def show_driver_orders_dialog(self):
        """查看司机订单对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("查看司机订单")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="司机ID:").grid(row=0, column=0, sticky=tk.W, pady=5)
        driver_id = ttk.Entry(frame, width=30)
        driver_id.grid(row=0, column=1, pady=5)
        
        ttk.Label(frame, text="日期:").grid(row=1, column=0, sticky=tk.W, pady=5)
        date = ttk.Entry(frame, width=30)
        date.insert(0, datetime.now().strftime('%Y-%m-%d'))
        date.grid(row=1, column=1, pady=5)
        
        def submit():
            try:
                orders = self.dispatcher.get_driver_orders(
                    driver_id=int(driver_id.get()),
                    date=date.get()
                )
                
                self.log(f"\n司机 {driver_id.get()} 的订单 (共 {len(orders)} 个):", "info")
                for order in orders:
                    self.log(f"  - {order}")
                
                dialog.destroy()
            except ValueError:
                messagebox.showerror("错误", "请输入有效的司机ID")
            except Exception as e:
                self.log(f"✗ 查询失败: {e}", "error")
                messagebox.showerror("错误", f"查询失败: {e}")
        
        ttk.Button(frame, text="查询", command=submit).grid(row=2, column=0, columnspan=2, pady=20)
    
    def show_batch_dispatch_dialog(self):
        """批量派工对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("批量派工")
        dialog.geometry("600x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="输入派工信息 (每行格式: 司机ID,订单ID,日期,时间段):").pack(anchor=tk.W)
        
        text = scrolledtext.ScrolledText(frame, width=70, height=20)
        text.pack(fill=tk.BOTH, expand=True, pady=10)
        text.insert(tk.END, "# 示例:\n# 123,1001,2025-11-20,09:00-12:00\n# 124,1002,2025-11-20,13:00-17:00\n")
        
        def submit():
            try:
                lines = text.get("1.0", tk.END).strip().split('\n')
                dispatch_list = []
                
                for line in lines:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    
                    parts = [p.strip() for p in line.split(',')]
                    if len(parts) >= 2:
                        dispatch_list.append({
                            'driver_id': int(parts[0]),
                            'order_id': int(parts[1]),
                            'date': parts[2] if len(parts) > 2 else None,
                            'time_slot': parts[3] if len(parts) > 3 else None
                        })
                
                if not dispatch_list:
                    messagebox.showwarning("提示", "没有有效的派工数据")
                    return
                
                self.log(f"\n开始批量派工 ({len(dispatch_list)} 个订单)...", "info")
                results = self.dispatcher.batch_dispatch(dispatch_list)
                
                success = sum(1 for r in results if r['result'].get('success'))
                self.log(f"✓ 批量派工完成: {success}/{len(results)} 成功", "success")
                
                messagebox.showinfo("完成", f"批量派工完成\n成功: {success}/{len(results)}")
                dialog.destroy()
                
            except Exception as e:
                self.log(f"✗ 批量派工失败: {e}", "error")
                messagebox.showerror("错误", f"批量派工失败: {e}")
        
        ttk.Button(frame, text="提交", command=submit).pack(pady=10)
    
    def show_batch_withdraw_dialog(self):
        """批量退工对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("批量退工")
        dialog.geometry("500x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="输入订单ID (每行一个):").pack(anchor=tk.W)
        
        text = scrolledtext.ScrolledText(frame, width=60, height=15)
        text.pack(fill=tk.BOTH, expand=True, pady=10)
        text.insert(tk.END, "# 示例:\n# 1001\n# 1002\n# 1003\n")
        
        ttk.Label(frame, text="退工原因:").pack(anchor=tk.W, pady=(10, 0))
        reason = ttk.Entry(frame, width=60)
        reason.pack(fill=tk.X, pady=5)
        
        def submit():
            try:
                lines = text.get("1.0", tk.END).strip().split('\n')
                order_ids = []
                
                for line in lines:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    order_ids.append(int(line))
                
                if not order_ids:
                    messagebox.showwarning("提示", "没有有效的订单ID")
                    return
                
                self.log(f"\n开始批量退工 ({len(order_ids)} 个订单)...", "info")
                results = self.dispatcher.batch_withdraw(order_ids, reason.get())
                
                success = sum(1 for r in results if r['result'].get('success'))
                self.log(f"✓ 批量退工完成: {success}/{len(results)} 成功", "success")
                
                messagebox.showinfo("完成", f"批量退工完成\n成功: {success}/{len(results)}")
                dialog.destroy()
                
            except Exception as e:
                self.log(f"✗ 批量退工失败: {e}", "error")
                messagebox.showerror("错误", f"批量退工失败: {e}")
        
        ttk.Button(frame, text="提交", command=submit).pack(pady=10)
    
    def view_logs(self):
        """查看日志"""
        try:
            if os.path.exists(config.LOG_FILE):
                with open(config.LOG_FILE, 'r', encoding='utf-8') as f:
                    logs = f.readlines()
                
                # 显示最后100行
                self.log("\n" + "=" * 60)
                self.log("最近的日志记录:", "info")
                self.log("=" * 60)
                for line in logs[-100:]:
                    self.log(line.strip())
            else:
                self.log("日志文件不存在", "warning")
        except Exception as e:
            self.log(f"✗ 读取日志失败: {e}", "error")
    
    def clear_output(self):
        """清空输出"""
        self.output_text.delete("1.0", tk.END)
        self.log("输出已清空", "info")
    
    def show_about(self):
        """显示关于信息"""
        about_text = """RPA调度系统自动化助手 v2.0

功能特性:
✓ 司机数据爬取（基本信息+详细资料）
✓ 工作排班分析（从路线数据提取）
✓ 路线数据爬取（支持多日查询）
✓ Excel报表导出（多工作表）
✓ JSON数据导出
✓ 数据统计分析

API端点:
• 司机: /drivers
• 路线: /routes
• 司机详情: /drivers/{id}

数据目录: data/

使用建议:
1. 快速测试 - 验证功能
2. 完整爬取 - 获取当天完整数据
3. 多日爬取 - 分析历史排班

注意: Token有效期24小时，需每日更新"""
        
        messagebox.showinfo("关于 RPA调度助手", about_text)


def main():
    """主函数"""
    root = tk.Tk()
    app = RPAAutomationGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
