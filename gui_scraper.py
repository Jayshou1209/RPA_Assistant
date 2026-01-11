"""
RPA调度系统 - 数据爬取工具
专注于数据采集和导出功能
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import os
import re
from datetime import datetime, timedelta
from api_client import APIClient
from scraper import DataScraper
import config
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(config.LOG_FILE, encoding='utf-8'),
    ]
)
logger = logging.getLogger(__name__)


class DataScraperGUI:
    """数据爬取工具GUI"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("RPA数据爬取工具 v1.0")
        self.root.geometry("1000x700")
        self.root.resizable(True, True)
        
        # 初始化变量
        self.api_client = None
        self.scraper = None
        self.enhanced_scraper = None
        self.real_scraper = None
        self.token_var = tk.StringVar(value=config.BEARER_TOKEN)
        self.status_var = tk.StringVar(value="就绪")
        self.last_data = None
        
        # 创建界面
        self.create_widgets()
        
        # 初始化API客户端
        self.initialize_client()
    
    def create_widgets(self):
        """创建界面组件"""
        # 主容器
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # 1. 功能按钮区域
        self.create_function_buttons(main_frame)
        
        # 2. 输出显示区域
        self.create_output_section(main_frame)
        
        # 3. 状态栏
        self.create_status_bar(main_frame)
    
    def create_token_section(self, parent):
        """创建Token管理区域"""
        token_frame = ttk.LabelFrame(parent, text="📝 Token管理", padding="10")
        token_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Token输入
        ttk.Label(token_frame, text="Bearer Token:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        token_entry = ttk.Entry(token_frame, textvariable=self.token_var, width=80)
        token_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        
        # 按钮组
        btn_frame = ttk.Frame(token_frame)
        btn_frame.grid(row=0, column=2, padx=(5, 0))
        
        ttk.Button(btn_frame, text="💾 保存Token", command=self.save_token).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🔄 重新加载", command=self.reload_token).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="✓ 测试连接", command=self.test_connection).pack(side=tk.LEFT, padx=2)
        
        token_frame.columnconfigure(1, weight=1)
    
    def create_function_buttons(self, parent):
        """创建功能按钮区域"""
        # 左侧按钮面板
        btn_frame = ttk.Frame(parent)
        btn_frame.grid(row=0, column=0, sticky=(tk.N, tk.W, tk.E), padx=(0, 10))
        
        # 数据爬取
        data_frame = ttk.LabelFrame(btn_frame, text="📊 数据爬取", padding="10")
        data_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(data_frame, text="👤 爬取司机数据", command=self.scrape_drivers_only, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="📅 爬取排班数据", command=self.scrape_schedules_only, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="📦 爬取订单数据", command=self.scrape_orders_only, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="💰 生成账单", command=self.generate_billing, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(data_frame, text="🚀 快速测试（10条）", command=self.quick_test_scrape, width=25).pack(fill=tk.X, pady=2)
        
        # 导出功能
        export_frame = ttk.LabelFrame(btn_frame, text="💾 导出与工具", padding="10")
        export_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(export_frame, text="📂 打开数据目录", command=self.open_data_folder, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(export_frame, text="🧹 清理数据目录", command=self.clean_data_folder, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(export_frame, text=" 导出为Excel", command=self.export_excel, width=25).pack(fill=tk.X, pady=2)
        
        # 系统操作
        system_frame = ttk.LabelFrame(btn_frame, text="⚙️ 系统", padding="10")
        system_frame.pack(fill=tk.X)
        
        ttk.Button(system_frame, text="📜 查看日志文件", command=self.view_logs, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(system_frame, text="🗑️ 清空输出", command=self.clear_output, width=25).pack(fill=tk.X, pady=2)
        ttk.Button(system_frame, text="ℹ️ 关于", command=self.show_about, width=25).pack(fill=tk.X, pady=2)
    
    def create_output_section(self, parent):
        """创建输出显示区域"""
        output_frame = ttk.LabelFrame(parent, text="📺 输出信息", padding="10")
        output_frame.grid(row=0, column=1, rowspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 创建带滚动条的文本框
        self.output_text = scrolledtext.ScrolledText(
            output_frame, 
            wrap=tk.WORD, 
            width=80, 
            height=30,
            font=("Consolas", 10)
        )
        self.output_text.pack(fill=tk.BOTH, expand=True)
        
        # 配置标签样式
        self.output_text.tag_config("success", foreground="green", font=("Consolas", 10, "bold"))
        self.output_text.tag_config("error", foreground="red", font=("Consolas", 10, "bold"))
        self.output_text.tag_config("info", foreground="blue")
        self.output_text.tag_config("warning", foreground="orange")
        
        self.log("=" * 60)
        self.log("欢迎使用 RPA数据爬取工具", "info")
        self.log("=" * 60)
    
    def create_status_bar(self, parent):
        """创建状态栏"""
        status_frame = ttk.Frame(parent)
        status_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Label(status_frame, text="状态:").pack(side=tk.LEFT)
        status_label = ttk.Label(status_frame, textvariable=self.status_var, foreground="blue")
        status_label.pack(side=tk.LEFT, padx=5)
        
        # 添加时间显示
        self.time_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ttk.Label(status_frame, textvariable=self.time_var).pack(side=tk.RIGHT)
        self.update_time()
    
    def update_time(self):
        """更新时间显示"""
        self.time_var.set(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.root.after(1000, self.update_time)
    
    def log(self, message, tag=""):
        """输出日志"""
        self.output_text.insert(tk.END, f"{message}\n", tag)
        self.output_text.see(tk.END)
        self.root.update_idletasks()
    
    def set_status(self, status):
        """设置状态"""
        self.status_var.set(status)
        self.root.update_idletasks()
    
    def initialize_client(self):
        """初始化API客户端"""
        try:
            from enhanced_scraper import EnhancedScraper
            from real_api_scraper import RealAPIScraper
            
            token = self.token_var.get()
            if not token:
                self.log("⚠️ Token为空，请先配置Token", "warning")
                return
            
            self.api_client = APIClient(token)
            self.scraper = DataScraper(self.api_client)
            self.enhanced_scraper = EnhancedScraper(self.api_client)
            self.real_scraper = RealAPIScraper(self.api_client)
            self.log("✓ API客户端初始化成功", "success")
        except ImportError as e:
            self.log(f"⚠️ 导入模块失败: {str(e)}", "warning")
            self.log("部分功能可能不可用", "warning")
            logger.error(f"导入失败: {e}", exc_info=True)
        except Exception as e:
            self.log(f"✗ 初始化失败: {str(e)}", "error")
            logger.error(f"初始化失败: {e}", exc_info=True)
    
    # ==================== Token管理 ====================
    
    def save_token(self):
        """保存Token"""
        try:
            new_token = self.token_var.get().strip()
            if not new_token:
                messagebox.showerror("错误", "Token不能为空")
                return
            
            # 保存到文件
            with open("token.txt", 'w', encoding='utf-8') as f:
                f.write(new_token)
            
            # 更新API客户端
            self.api_client.update_token(new_token)
            
            self.log("✓ Token已保存并更新", "success")
            messagebox.showinfo("成功", "Token已保存")
        except Exception as e:
            self.log(f"✗ 保存Token失败: {str(e)}", "error")
            messagebox.showerror("错误", f"保存失败: {str(e)}")
    
    def reload_token(self):
        """重新加载Token"""
        try:
            with open("token.txt", 'r', encoding='utf-8-sig') as f:
                token = f.read().strip()
            self.token_var.set(token)
            self.api_client.update_token(token)
            self.log("✓ Token已重新加载", "success")
        except Exception as e:
            self.log(f"✗ 加载Token失败: {str(e)}", "error")
            messagebox.showerror("错误", f"加载失败: {str(e)}")
    
    def test_connection(self):
        """测试API连接"""
        def test():
            try:
                self.set_status("正在测试连接...")
                self.log("开始测试API连接...")
                
                # 使用verify_connection方法
                success, message = self.api_client.verify_connection()
                
                if success:
                    self.log("✓ API连接测试成功", "success")
                    self.log(f"  {message}", "info")
                    self.set_status("连接正常")
                    messagebox.showinfo("成功", f"API连接正常\n\n{message}")
                else:
                    self.log("✗ 连接测试失败", "error")
                    self.log(f"  {message}", "error")
                    self.set_status("连接失败")
                    messagebox.showerror("失败", f"连接失败\n\n{message}")
            except Exception as e:
                self.log(f"✗ 连接测试出错: {str(e)}", "error")
                self.set_status("连接出错")
                messagebox.showerror("错误", f"测试出错: {str(e)}")
        
        threading.Thread(target=test, daemon=True).start()
    
    # ==================== 数据爬取 ====================
    
    def scrape_drivers_only(self):
        """仅爬取司机数据（包含完整的详细信息）"""
        def scrape():
            try:
                self.set_status("正在爬取司机数据...")
                self.log("\n" + "="*60)
                self.log("开始爬取司机完整数据...", "info")
                self.log("提示：将获取每位司机的详细证件信息和车辆信息", "info")
                self.log("-"*60)
                
                # 使用新的详细爬取方法
                drivers = self.real_scraper.get_all_drivers_with_full_details(
                    per_page=100,
                    progress_callback=lambda current, total, msg: self.log(f"  进度: {current}/{total} - {msg}", "info")
                )
                
                self.log("-"*60)
                self.log(f"✓ 成功获取 {len(drivers)} 位司机的完整数据", "success")
                self.last_data = {"drivers": drivers}
                
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_file = os.path.join(config.DATA_DIR, f"司机完整数据_{timestamp}.xlsx")
                
                # 直接导出Excel
                self.log("\n开始导出Excel...", "info")
                result = self.export_drivers_excel(drivers, excel_file)
                if result:
                    self.log(result, "success")
                    # 自动打开文件
                    try:
                        os.startfile(excel_file)
                        self.log(f"✓ 已自动打开Excel文件", "success")
                    except Exception as e:
                        self.log(f"无法自动打开文件: {e}", "warning")
                
                self.set_status("爬取完成")
                
            except Exception as e:
                self.log(f"✗ 爬取出错: {str(e)}", "error")
                self.set_status("出错")
                logger.error(f"爬取司机数据出错: {e}", exc_info=True)
        
        threading.Thread(target=scrape, daemon=True).start()
    
    def scrape_schedules_only(self):
        """仅爬取排班数据（多线程并发）"""
        def scrape():
            try:
                self.set_status("正在爬取排班数据...")
                self.log("\n" + "="*60)
                self.log("开始爬取排班数据...", "info")
                
                date = datetime.now().strftime('%Y-%m-%d')
                routes = self.real_scraper.get_all_routes(date=date, per_page=100)
                
                self.log(f"✓ 成功获取 {len(routes)} 条排班数据", "success")
                self.last_data = {"routes": routes, "date": date}
                
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_file = os.path.join(config.DATA_DIR, f"排班数据_{date}_{timestamp}.xlsx")
                
                # 直接导出Excel
                self.log("开始导出Excel...", "info")
                result = self.export_schedules_excel(routes, excel_file, date)
                if result:
                    self.log(result, "success")
                
                self.set_status("爬取完成")
                
            except Exception as e:
                self.log(f"✗ 爬取出错: {str(e)}", "error")
                self.set_status("出错")
                logger.error(f"爬取排班数据出错: {e}", exc_info=True)
        
        threading.Thread(target=scrape, daemon=True).start()
    
    def scrape_orders_only(self):
        """仅爬取订单数据（多线程并发）"""
        def scrape():
            try:
                self.set_status("正在爬取订单数据...")
                self.log("\n" + "="*60)
                self.log("开始爬取订单数据...", "info")
                
                date = datetime.now().strftime('%Y-%m-%d')
                rides = self.real_scraper.get_all_rides(date=date, per_page=500)
                
                self.log(f"✓ 成功获取 {len(rides)} 条订单数据", "success")
                
                # 获取每个订单的详细信息（并发处理，大幅提速）
                self.log("\n获取订单详细信息（价格、Co Pay、TOLL等）- 并发处理中...", "info")
                detailed_rides = []
                
                def fetch_ride_detail(ride):
                    """获取单个订单详情"""
                    try:
                        ride_id = ride.get('id')
                        ride_detail = self.api_client.get(f'/fleet/rides/{ride_id}')
                        
                        # 提取价格信息
                        vendor_amount = float(ride_detail.get('vendor_amount', 0) or 0)
                        original_price = vendor_amount
                        
                        # 从notes中提取co_pay
                        notes = ride_detail.get('notes', [])
                        co_pay = 0
                        for note in notes:
                            note_text = note.get('note', '')
                            match = re.search(r'Co[- ]?Pay[:\s]*\$?([\d.]+)', note_text, re.IGNORECASE)
                            if match:
                                co_pay = float(match.group(1))
                                break
                        
                        # 计算订单价格和toll
                        order_status = ride_detail.get('status', '')
                        if order_status in ['no_show', 'driver_canceled']:
                            order_price = 0
                            co_pay = 0
                            toll_fee = 0
                        else:
                            if co_pay > 0:
                                order_price = round(vendor_amount - co_pay, 2)
                                toll_fee = 0
                                original_price = vendor_amount
                            else:
                                order_price = round(original_price - co_pay, 2)
                                toll_fee = round(vendor_amount - order_price, 2)
                        
                        # 合并详细信息
                        ride['vendor_amount'] = vendor_amount
                        ride['original_price'] = original_price
                        ride['co_pay'] = co_pay
                        ride['order_price'] = order_price
                        ride['toll_fee'] = toll_fee
                        ride['distance'] = ride_detail.get('distance', 0)
                        ride['pickup_at'] = ride_detail.get('pickup_at', ride.get('schedule_time', ''))
                        ride['start_address'] = ride_detail.get('start_address', ride.get('pickup_address', ''))
                        ride['destination_address'] = ride_detail.get('destination_address', ride.get('dropoff_address', ''))
                        ride['first_name'] = ride_detail.get('first_name', '')
                        ride['last_name'] = ride_detail.get('last_name', '')
                        
                        return ride
                    except Exception as e:
                        # 失败时返回基本信息
                        ride['vendor_amount'] = 0
                        ride['co_pay'] = 0
                        ride['order_price'] = 0
                        ride['toll_fee'] = 0
                        return ride
                
                # 使用线程池并发处理（10个线程同时处理，速度提升10倍）
                with ThreadPoolExecutor(max_workers=10) as executor:
                    futures = {executor.submit(fetch_ride_detail, ride): ride for ride in rides}
                    completed = 0
                    for future in as_completed(futures):
                        detailed_ride = future.result()
                        detailed_rides.append(detailed_ride)
                        completed += 1
                        if completed % 20 == 0 or completed == len(rides):
                            self.log(f"  进度: {completed}/{len(rides)} 条订单", "info")
                
                self.log(f"✓ 已获取 {len(detailed_rides)} 条订单详细信息", "success")
                
                self.last_data = {"rides": detailed_rides, "date": date}
                
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_file = os.path.join(config.DATA_DIR, f"订单数据_{date}_{timestamp}.xlsx")
                
                # 直接导出Excel
                self.log("开始导出Excel...", "info")
                result = self.export_orders_excel(detailed_rides, excel_file, date)
                if result:
                    self.log(result, "success")
                
                self.set_status("爬取完成")
                
            except Exception as e:
                self.log(f"✗ 爬取出错: {str(e)}", "error")
                self.set_status("出错")
        
        threading.Thread(target=scrape, daemon=True).start()
    
    def generate_billing(self):
        """生成账单 - 支持日期范围"""
        # 创建日期选择对话框
        date_dialog = tk.Toplevel(self.root)
        date_dialog.title("选择账单日期范围")
        date_dialog.geometry("400x250")
        date_dialog.transient(self.root)
        date_dialog.grab_set()
        
        # 居中显示
        date_dialog.update_idletasks()
        x = (date_dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (date_dialog.winfo_screenheight() // 2) - (250 // 2)
        date_dialog.geometry(f"400x250+{x}+{y}")
        
        ttk.Label(date_dialog, text="选择账单日期范围:", font=("Arial", 11, "bold")).pack(pady=15)
        
        # 日期输入框
        date_frame = ttk.Frame(date_dialog)
        date_frame.pack(pady=10)
        
        ttk.Label(date_frame, text="开始日期:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        start_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        start_date_entry = ttk.Entry(date_frame, textvariable=start_date_var, width=15)
        start_date_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(date_frame, text="结束日期:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        end_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        end_date_entry = ttk.Entry(date_frame, textvariable=end_date_var, width=15)
        end_date_entry.grid(row=1, column=1, padx=5, pady=5)
        
        # 快捷按钮
        quick_frame = ttk.Frame(date_dialog)
        quick_frame.pack(pady=10)
        
        def set_yesterday():
            yesterday = datetime.now() - timedelta(days=1)
            yesterday_str = yesterday.strftime('%Y-%m-%d')
            start_date_var.set(yesterday_str)
            end_date_var.set(yesterday_str)
            start_date_entry.update()
            end_date_entry.update()
        
        def set_this_week():
            today = datetime.now()
            start = today - timedelta(days=today.weekday())
            end = start + timedelta(days=6)
            start_date_var.set(start.strftime('%Y-%m-%d'))
            end_date_var.set(end.strftime('%Y-%m-%d'))
            start_date_entry.update()
            end_date_entry.update()
        
        def set_last_week():
            today = datetime.now()
            start = today - timedelta(days=today.weekday() + 7)
            end = start + timedelta(days=6)
            start_date_var.set(start.strftime('%Y-%m-%d'))
            end_date_var.set(end.strftime('%Y-%m-%d'))
            start_date_entry.update()
            end_date_entry.update()
        
        ttk.Button(quick_frame, text="昨天", command=set_yesterday).pack(side=tk.LEFT, padx=3)
        ttk.Button(quick_frame, text="本周", command=set_this_week).pack(side=tk.LEFT, padx=3)
        ttk.Button(quick_frame, text="上周", command=set_last_week).pack(side=tk.LEFT, padx=3)
        
        def start_generate():
            start_date = start_date_var.get()
            end_date = end_date_var.get()
            date_dialog.destroy()
            self._generate_billing_for_range(start_date, end_date)
        
        ttk.Button(date_dialog, text="生成账单", command=start_generate, width=20).pack(pady=15)
    
    def _generate_billing_for_range(self, start_date, end_date):
        """生成指定日期范围的账单"""
        def task():
            try:
                self.set_status(f"正在生成 {start_date} 至 {end_date} 的账单...")
                self.log("=" * 60)
                self.log(f"开始生成账单: {start_date} 至 {end_date}", "info")
                self.log("=" * 60)
                
                # 解析日期
                from datetime import datetime, timedelta
                start = datetime.strptime(start_date, '%Y-%m-%d')
                end = datetime.strptime(end_date, '%Y-%m-%d')
                
                # 计算日期范围
                delta = end - start
                if delta.days < 0:
                    self.log("✗ 结束日期必须大于等于开始日期", "error")
                    messagebox.showerror("错误", "结束日期必须大于等于开始日期")
                    self.set_status("就绪")
                    return
                
                # 收集所有日期的订单
                all_rides = []
                current = start
                
                self.log(f"\n📅 需要处理 {delta.days + 1} 天的数据", "info")
                
                while current <= end:
                    date_str = current.strftime('%Y-%m-%d')
                    self.log(f"\n获取 {date_str} 的订单...", "info")
                    
                    try:
                        rides = self.real_scraper.get_all_rides(
                            date=date_str,
                            per_page=500,
                            statuses='finished,no_show,driver_canceled'
                        )
                        self.log(f"  ✓ {date_str}: {len(rides)} 条订单", "success")
                        all_rides.extend(rides)
                    except Exception as e:
                        self.log(f"  ✗ {date_str}: 获取失败 - {e}", "error")
                    
                    current += timedelta(days=1)
                
                self.log(f"\n✓ 总共获取 {len(all_rides)} 条订单", "success")
                
                if len(all_rides) == 0:
                    self.log(f"\n⚠️ 未找到符合条件的订单", "warning")
                    messagebox.showwarning("提示", "未找到符合条件的订单")
                    self.set_status("就绪")
                    return
                
                # 获取订单详细信息（并发处理，大幅提速）
                self.log("\n2️⃣ 获取订单详细信息（价格、Co Pay、TOLL）- 并发处理中...", "info")
                detailed_rides = []
                
                def fetch_billing_ride_detail(ride):
                    """获取单个账单订单详情"""
                    try:
                        ride_id = ride.get('id')
                        detail = self.api_client.get(f'/fleet/rides/{ride_id}')
                        ride_detail = detail.get('ride', {})
                        
                        # 提取价格信息
                        vendor_amount = float(ride_detail.get('vendor_amount', 0) or 0)
                        
                        # 从events中提取notes里的价格（"reserved the ride for $XX.XX"）
                        events = ride_detail.get('events', [])
                        notes_price = 0
                        for event in events:
                            body = event.get('body', '')
                            match = re.search(r'reserved.*for\s+\$([0-9]+\.?[0-9]*)', body, re.IGNORECASE)
                            if match:
                                notes_price = float(match.group(1))
                                break
                        
                        # 从notes中提取Co Pay（必须label中含$符号）
                        # 优先检查icon='private'的notes，其次检查description中含'collect'或'cash'的notes
                        notes = ride_detail.get('notes', [])
                        co_pay = 0
                        for note in notes:
                            label = note.get('label', '')
                            description = note.get('description', '').lower()
                            icon = note.get('icon', '')
                            
                            # 先在label中查找$符号和金额
                            match = re.search(r'\$([0-9]+\.?[0-9]*)', label)
                            if match:
                                # 如果找到了$金额，再检查条件
                                if icon == 'private' or 'collect' in description or 'cash' in description:
                                    co_pay = float(match.group(1))
                                    break
                        
                        # 检查订单状态
                        order_status = ride_detail.get('status', '')
                        
                        # no_show和driver_canceled订单特殊处理
                        if order_status in ['no_show', 'driver_canceled']:
                            order_price = 0
                            toll_fee = 0
                            original_price = 0
                            co_pay = 0
                        else:
                            # 如果notes里有价格，订单价格 = notes价格 - co_pay
                            if notes_price > 0:
                                order_price = round(notes_price - co_pay, 2)
                                toll_fee = round(vendor_amount - notes_price + co_pay, 2)
                                original_price = notes_price
                            else:
                                # 如果notes里没有价格，订单价格 = vendor_amount
                                order_price = vendor_amount
                                toll_fee = 0
                                original_price = vendor_amount
                        
                        # 合并所有信息到ride
                        ride['vendor_amount'] = vendor_amount
                        ride['original_price'] = original_price
                        ride['co_pay'] = co_pay
                        ride['order_price'] = order_price
                        ride['toll_fee'] = toll_fee
                        ride['has_notes_price'] = notes_price > 0  # 标记是否有events价格
                        ride['distance'] = ride_detail.get('distance', 0)
                        ride['pickup_at'] = ride_detail.get('pickup_at', ride.get('schedule_time', ''))
                        ride['start_address'] = ride_detail.get('start_address', ride.get('pickup_address', ''))
                        ride['destination_address'] = ride_detail.get('destination_address', ride.get('dropoff_address', ''))
                        ride['first_name'] = ride_detail.get('first_name', '')
                        ride['last_name'] = ride_detail.get('last_name', '')
                        
                        return ride
                    except Exception as e:
                        # 失败时返回基本信息
                        ride['vendor_amount'] = 0
                        ride['co_pay'] = 0
                        ride['order_price'] = 0
                        ride['toll_fee'] = 0
                        return ride
                
                # 使用线程池并发处理（15个线程同时处理，速度提升15倍）
                with ThreadPoolExecutor(max_workers=15) as executor:
                    futures = {executor.submit(fetch_billing_ride_detail, ride): ride for ride in all_rides}
                    completed = 0
                    for future in as_completed(futures):
                        detailed_ride = future.result()
                        detailed_rides.append(detailed_ride)
                        completed += 1
                        if completed % 30 == 0 or completed == len(all_rides):
                            self.log(f"  进度: {completed}/{len(all_rides)} 条订单", "info")
                
                self.log(f"✓ 已获取 {len(detailed_rides)} 条订单详情", "success")
                
                # 按司机分组统计
                self.log("\n💰 开始生成账单统计...", "info")
                driver_billing = {}
                
                for ride in detailed_rides:
                    driver_id = ride.get('driver_id')
                    if not driver_id:
                        continue
                    
                    status = ride.get('status', '')
                    
                    if driver_id not in driver_billing:
                        driver_billing[driver_id] = {
                            'driver_id': driver_id,
                            'driver_name': f"{ride.get('driver_first_name', '')} {ride.get('driver_last_name', '')}".strip(),
                            'finished_count': 0,  # 只统计finished状态的订单
                            'no_show': 0,
                            'driver_canceled': 0,
                            'total_amount': 0,
                            'rides': []
                        }
                    
                    # 只有finished状态才计入完成订单数
                    if status == 'finished':
                        driver_billing[driver_id]['finished_count'] += 1
                    elif status == 'no_show':
                        driver_billing[driver_id]['no_show'] += 1
                    elif status == 'driver_canceled':
                        driver_billing[driver_id]['driver_canceled'] += 1
                    
                    # 简化版价格计算
                    if status in ['no_show', 'driver_canceled']:
                        amount = 5.0
                    else:
                        amount = float(ride.get('vendor_amount', 0) or 0)
                    
                    driver_billing[driver_id]['total_amount'] += amount
                    driver_billing[driver_id]['rides'].append(ride)
                
                # 输出账单摘要
                self.log("\n" + "=" * 60)
                self.log("📊 账单摘要", "info")
                self.log("=" * 60)
                
                for driver_id, billing in driver_billing.items():
                    self.log(f"\n司机: {billing['driver_name']} (ID: {driver_id})")
                    self.log(f"  完成订单: {billing['finished_count']} 条")
                    self.log(f"  No Show: {billing['no_show']} 条 | Driver Canceled: {billing['driver_canceled']} 条", "warning")
                    self.log(f"  总金额: ${billing['total_amount']:.2f}", "success")
                
                # 保存数据
                self.last_data = {
                    'start_date': start_date,
                    'end_date': end_date,
                    'billing': list(driver_billing.values()),
                    'all_rides': all_rides
                }
                
                self.log("\n" + "=" * 60)
                self.log("✓ 账单生成完成！", "success")
                self.log(f"日期范围: {start_date} 至 {end_date}", "info")
                self.log(f"司机数: {len(driver_billing)} 位", "info")
                self.log(f"订单数: {len(all_rides)} 条", "info")
                self.log("=" * 60)
                
                # 自动导出为Excel
                self.log("\n正在自动导出Excel...", "info")
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_file = os.path.join(config.DATA_DIR, f"账单_{start_date}_至_{end_date}_{timestamp}.xlsx")
                
                try:
                    # 使用export_excel的逻辑直接导出
                    import pandas as pd
                    from openpyxl.styles import Font, PatternFill, numbers
                    
                    billing_data = list(driver_billing.values())
                    
                    # 准备导出数据（按司机分组，每个司机包含详细订单+汇总行）
                    export_rows = []
                    summary_row_indices = []  # 记录汇总行的索引
                    no_notes_price_indices = []  # 记录没有events价格的订单行索引
                    
                    for billing in billing_data:
                        driver_name = billing.get('driver_name', '')
                        rides = billing.get('rides', [])
                        
                        # 添加该司机的所有订单
                        for ride in rides:
                            status = ride.get('status', '')
                            co_pay = float(ride.get('co_pay', 0) or 0)
                            order_price = float(ride.get('order_price', 0) or 0)
                            toll_fee = float(ride.get('toll_fee', 0) or 0)
                            has_notes_price = ride.get('has_notes_price', True)
                            
                            # 如果没有events价格，记录行索引
                            if not has_notes_price:
                                no_notes_price_indices.append(len(export_rows))
                            
                            # 计算NO SHOW金额
                            no_show = 5.0 if status in ['no_show', 'driver_canceled'] else 0.0
                            
                            export_rows.append({
                                '司机姓名': driver_name,
                                '订单数': None,
                                '总收入': None,
                                '订单ID': ride.get('id', ''),
                                '接客时间': ride.get('pickup_at', ride.get('schedule_time', '')),
                                '接客地点': ride.get('start_address', ride.get('pickup_address', '')),
                                '送达地点': ride.get('destination_address', ride.get('dropoff_address', '')),
                                '乘客姓名': f"{ride.get('first_name', '')} {ride.get('last_name', '')}".strip() or ride.get('customer_name', ''),
                                '订单价格': order_price,
                                'NO SHOW': no_show,
                                'Co Pay': co_pay,
                                'TOLL': toll_fee,
                                '状态': status
                            })
                        
                        # 添加该司机的汇总行
                        finished_count = billing.get('finished_count', 0)
                        summary_row_indices.append(len(export_rows))
                        export_rows.append({
                            '司机姓名': driver_name,
                            '订单数': finished_count,
                            '总收入': None,
                            '订单ID': '',
                            '接客时间': '',
                            '接客地点': '',
                            '送达地点': '',
                            '乘客姓名': '',
                            '订单价格': None,
                            'NO SHOW': None,
                            'Co Pay': None,
                            'TOLL': None,
                            '状态': ''
                        })
                    
                    # 添加底部总计行
                    total_finished_count = sum(billing.get('finished_count', 0) for billing in billing_data)
                    total_row_index = len(export_rows)
                    export_rows.append({
                        '司机姓名': '总计',
                        '订单数': total_finished_count,
                        '总收入': None,
                        '订单ID': '',
                        '接客时间': '',
                        '接客地点': '',
                        '送达地点': '',
                        '乘客姓名': '',
                        '订单价格': None,
                        'NO SHOW': None,
                        'Co Pay': None,
                        'TOLL': None,
                        '状态': ''
                    })
                    
                    # 创建DataFrame并导出
                    df = pd.DataFrame(export_rows)
                    
                    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='账单详情', index=False)
                        workbook = writer.book
                        worksheet = writer.sheets['账单详情']
                        
                        # 填充公式和样式
                        for idx, row_idx in enumerate(summary_row_indices):
                            excel_row = row_idx + 2
                            start_row = 2 if idx == 0 else summary_row_indices[idx-1] + 3
                            
                            worksheet.cell(row=excel_row, column=9, value=f'=SUM(I{start_row}:I{excel_row-1})')
                            worksheet.cell(row=excel_row, column=10, value=f'=SUM(J{start_row}:J{excel_row-1})')
                            worksheet.cell(row=excel_row, column=11, value=f'=SUM(K{start_row}:K{excel_row-1})')
                            worksheet.cell(row=excel_row, column=12, value=f'=SUM(L{start_row}:L{excel_row-1})')
                            worksheet.cell(row=excel_row, column=3, value=f'=I{excel_row}+J{excel_row}+K{excel_row}+L{excel_row}')
                        
                        last_row = total_row_index + 2
                        summary_rows_excel = [str(idx + 2) for idx in summary_row_indices]
                        worksheet.cell(row=last_row, column=9, value=f'={"+".join([f"I{row}" for row in summary_rows_excel])}')
                        worksheet.cell(row=last_row, column=10, value=f'={"+".join([f"J{row}" for row in summary_rows_excel])}')
                        worksheet.cell(row=last_row, column=11, value=f'={"+".join([f"K{row}" for row in summary_rows_excel])}')
                        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色
                        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色
                        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # 橙色
                        bold_font = Font(bold=True)
                        
                        # 标记没有events价格的订单行（绿色填充）
                        for row_idx in no_notes_price_indices:
                            for col in range(1, 14):
                                cell = worksheet.cell(row=row_idx + 2, column=col)
                                cell.fill = green_fill
                        
                        # 标记汇总行（黄色填充）
                        for row_idx in summary_row_indices:
                            for col in range(1, 14):
                                cell = worksheet.cell(row=row_idx + 2, column=col)
                                cell.fill = yellow_fill
                                cell.font = bold_font
                        
                        # 标记总计行（橙色填充）
                        for row_idx in summary_row_indices:
                            for col in range(1, 14):
                                cell = worksheet.cell(row=row_idx + 2, column=col)
                                cell.fill = yellow_fill
                                cell.font = bold_font
                        
                        for col in range(1, 14):
                            cell = worksheet.cell(row=last_row, column=col)
                            cell.fill = orange_fill
                            cell.font = bold_font
                        
                        # 货币格式
                        for row in range(2, last_row + 1):
                            for col in [3, 9, 10, 11, 12]:
                                worksheet.cell(row=row, column=col).number_format = '$#,##0.00'
                        
                        # 列宽
                        worksheet.column_dimensions['A'].width = 20
                        worksheet.column_dimensions['C'].width = 15
                        worksheet.column_dimensions['E'].width = 20
                        worksheet.column_dimensions['F'].width = 40
                        worksheet.column_dimensions['G'].width = 40
                    
                    self.log(f"✓ Excel已导出: {excel_file}", "success")
                    self.set_status("就绪")
                    messagebox.showinfo("完成", 
                        f"账单生成并导出成功！\n\n"
                        f"日期范围: {start_date} 至 {end_date}\n"
                        f"司机: {len(driver_billing)} 位\n"
                        f"订单: {len(all_rides)} 条\n\n"
                        f"文件已保存:\n{excel_file}")
                        
                except Exception as export_error:
                    import traceback
                    self.log(f"✗ Excel导出失败: {export_error}", "error")
                    self.log(traceback.format_exc(), "error")
                    self.set_status("就绪")
                    messagebox.showwarning("部分完成", 
                        f"账单生成完成，但导出Excel失败！\n\n"
                        f"日期范围: {start_date} 至 {end_date}\n"
                        f"司机: {len(driver_billing)} 位\n"
                        f"订单: {len(all_rides)} 条\n\n"
                        f"错误: {export_error}\n\n"
                        f"请手动点击'导出为Excel'按钮")
                
            except Exception as e:
                import traceback
                self.log(f"✗ 生成账单失败: {e}", "error")
                self.log(traceback.format_exc(), "error")
                self.set_status("就绪")
                messagebox.showerror("错误", f"生成账单失败:\n{e}")
        
        threading.Thread(target=task, daemon=True).start()
    
    def quick_test_scrape(self):
        """快速测试爬取10条数据"""
        def test():
            try:
                self.set_status("正在快速测试...")
                self.log("\n" + "="*60)
                self.log("快速测试：获取前10条司机数据", "info")
                
                # 使用real_scraper获取司机数据
                drivers = self.real_scraper.get_all_drivers(per_page=10)
                
                self.log(f"✓ 成功获取 {len(drivers)} 条测试数据", "success")
                for i, driver in enumerate(drivers[:10], 1):
                    name = f"{driver.get('first_name', '')} {driver.get('last_name', '')}".strip()
                    phone = driver.get('phone', 'N/A')
                    self.log(f"  {i}. {name} - {phone}")
                self.set_status("测试完成")
                
            except Exception as e:
                self.log(f"✗ 测试出错: {str(e)}", "error")
                self.set_status("出错")
        
        threading.Thread(target=test, daemon=True).start()
    
    # ==================== 导出功能 ====================
    
    def open_data_folder(self):
        """打开数据目录"""
        try:
            data_dir = config.DATA_DIR
            if not os.path.exists(data_dir):
                os.makedirs(data_dir)
            os.startfile(data_dir)
            self.log("✓ 已打开数据目录", "success")
        except Exception as e:
            self.log(f"✗ 打开目录失败: {str(e)}", "error")
            messagebox.showerror("错误", f"打开失败: {str(e)}")
    
    def clean_data_folder(self):
        """清理数据目录"""
        if messagebox.askyesno("确认", "确定要清理数据目录吗？\n这将删除所有导出的文件。"):
            try:
                data_dir = config.DATA_DIR
                if os.path.exists(data_dir):
                    for file in os.listdir(data_dir):
                        if file.endswith(('.json', '.xlsx', '.csv')):
                            os.remove(os.path.join(data_dir, file))
                self.log("✓ 数据目录已清理", "success")
                messagebox.showinfo("成功", "数据目录已清理")
            except Exception as e:
                self.log(f"✗ 清理失败: {str(e)}", "error")
                messagebox.showerror("错误", f"清理失败: {str(e)}")
    
    def export_json(self):
        """导出为JSON"""
        if not self.last_data:
            messagebox.showwarning("警告", "没有可导出的数据，请先执行爬取操作")
            return
        
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")],
                initialdir=config.DATA_DIR,
                initialfile=f"data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            )
            
            if filename:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(self.last_data, f, ensure_ascii=False, indent=2)
                self.log(f"✓ 数据已导出到: {filename}", "success")
                messagebox.showinfo("成功", "数据导出成功")
        except Exception as e:
            self.log(f"✗ 导出失败: {str(e)}", "error")
            messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def export_excel(self):
        """导出为Excel（详细订单+司机汇总）"""
        if not self.last_data:
            messagebox.showwarning("警告", "没有可导出的数据，请先生成账单")
            return
        
        # 检查是否有账单数据
        if 'billing' not in self.last_data:
            messagebox.showwarning("警告", "请先生成账单后再导出")
            return
        
        # 检查pandas和openpyxl是否已安装
        try:
            import pandas as pd
            from openpyxl.styles import Font, PatternFill, numbers
        except ImportError as e:
            self.log(f"✗ 缺少必要的库: {e}", "error")
            messagebox.showerror("错误", "缺少必要的库\n\n请安装:\npip install pandas openpyxl")
            return
        
        try:
            billing_data = self.last_data.get('billing', [])
            start_date = self.last_data.get('start_date', '')
            end_date = self.last_data.get('end_date', '')
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                initialdir=config.DATA_DIR,
                initialfile=f"账单_{start_date}_至_{end_date}_{timestamp}.xlsx"
            )
            
            if not filename:
                return
            
            # 准备导出数据（按司机分组，每个司机包含详细订单+汇总行）
            export_rows = []
            summary_row_indices = []  # 记录汇总行的索引
            
            for billing in billing_data:
                driver_name = billing.get('driver_name', '')
                rides = billing.get('rides', [])
                
                # 添加该司机的所有订单（包括driver_canceled）
                for ride in rides:
                    status = ride.get('status', '')
                    co_pay = float(ride.get('co_pay', 0) or 0)
                    order_price = float(ride.get('order_price', 0) or 0)
                    toll_fee = float(ride.get('toll_fee', 0) or 0)
                    distance = float(ride.get('distance', 0) or 0)
                    
                    # 计算NO SHOW金额：no_show和driver_canceled订单为$5
                    no_show = 5.0 if status in ['no_show', 'driver_canceled'] else 0.0
                    
                    export_rows.append({
                        '司机姓名': driver_name,
                        '订单数': None,  # 详细行不显示订单数
                        '总收入': None,
                        '订单ID': ride.get('id', ''),
                        '接客时间': ride.get('pickup_at', ride.get('schedule_time', '')),
                        '接客地点': ride.get('start_address', ride.get('pickup_address', '')),
                        '送达地点': ride.get('destination_address', ride.get('dropoff_address', '')),
                        '乘客姓名': f"{ride.get('first_name', '')} {ride.get('last_name', '')}".strip() or ride.get('customer_name', ''),
                        '订单价格': order_price,
                        'NO SHOW': no_show,
                        'Co Pay': co_pay,
                        'TOLL': toll_fee,
                        '状态': status
                    })
                
                # 添加该司机的汇总行
                finished_count = billing.get('finished_count', 0)  # 只计算finished的订单数
                
                summary_row_indices.append(len(export_rows))  # 记录汇总行索引
                export_rows.append({
                    '司机姓名': driver_name,
                    '订单数': finished_count,  # 只显示finished订单数
                    '总收入': None,  # 将用公式填充
                    '订单ID': '',
                    '接客时间': '',
                    '接客地点': '',
                    '送达地点': '',
                    '乘客姓名': '',
                    '订单价格': None,  # 将用公式填充
                    'NO SHOW': None,  # 将用公式填充
                    'Co Pay': None,  # 将用公式填充
                    'TOLL': None,  # 将用公式填充
                    '状态': ''
                })
            
            # 添加底部总计行
            total_finished_count = sum(billing.get('finished_count', 0) for billing in billing_data)
            total_row_index = len(export_rows)
            
            export_rows.append({
                '司机姓名': '总计',
                '订单数': total_finished_count,
                '总收入': None,  # 将用公式填充
                '订单ID': '',
                '接客时间': '',
                '接客地点': '',
                '送达地点': '',
                '乘客姓名': '',
                '订单价格': None,  # 将用公式填充
                'NO SHOW': None,  # 将用公式填充
                'Co Pay': None,  # 将用公式填充
                'TOLL': None,  # 将用公式填充
                '状态': ''
            })
            
            # 创建DataFrame
            df = pd.DataFrame(export_rows)
            
            # 导出到Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='账单详情', index=False)
                
                # 获取工作表进行样式设置
                workbook = writer.book
                worksheet = writer.sheets['账单详情']
                
                # 填充司机汇总行的Excel公式
                for idx, row_idx in enumerate(summary_row_indices):
                    excel_row = row_idx + 2  # +2 因为Excel从1开始，且有表头
                    
                    # 找到该司机的订单起始行
                    if idx == 0:
                        start_row = 2
                    else:
                        start_row = summary_row_indices[idx-1] + 3  # 上一个汇总行的下一行
                    
                    # 订单价格合计 (I列)
                    worksheet.cell(row=excel_row, column=9, value=f'=SUM(I{start_row}:I{excel_row-1})')
                    # NO SHOW合计 (J列)
                    worksheet.cell(row=excel_row, column=10, value=f'=SUM(J{start_row}:J{excel_row-1})')
                    # Co Pay合计 (K列)
                    worksheet.cell(row=excel_row, column=11, value=f'=SUM(K{start_row}:K{excel_row-1})')
                    # TOLL合计 (L列)
                    worksheet.cell(row=excel_row, column=12, value=f'=SUM(L{start_row}:L{excel_row-1})')
                    # 总收入 = 订单价格 + Co Pay + TOLL + NO SHOW
                    worksheet.cell(row=excel_row, column=3, value=f'=I{excel_row}+J{excel_row}+K{excel_row}+L{excel_row}')
                
                # 填充底部总计行的Excel公式
                last_row = total_row_index + 2  # +2 因为Excel从1开始，且有表头
                
                # 总计行只汇总各司机的汇总行，而不是所有订单详细行
                # 构建只汇总司机汇总行的公式
                summary_rows_excel = [str(idx + 2) for idx in summary_row_indices]
                sum_formula_order_price = '+'.join([f'I{row}' for row in summary_rows_excel])
                sum_formula_no_show = '+'.join([f'J{row}' for row in summary_rows_excel])
                sum_formula_copay = '+'.join([f'K{row}' for row in summary_rows_excel])
                sum_formula_toll = '+'.join([f'L{row}' for row in summary_rows_excel])
                
                worksheet.cell(row=last_row, column=9, value=f'={sum_formula_order_price}')
                worksheet.cell(row=last_row, column=10, value=f'={sum_formula_no_show}')
                worksheet.cell(row=last_row, column=11, value=f'={sum_formula_copay}')
                worksheet.cell(row=last_row, column=12, value=f'={sum_formula_toll}')
                # 总收入 = 订单价格总计 + Co Pay总计 + TOLL总计 + NO SHOW总计
                worksheet.cell(row=last_row, column=3, value=f'=I{last_row}+J{last_row}+K{last_row}+L{last_row}')
                
                # 设置汇总行样式（黄色背景+加粗）
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                bold_font = Font(bold=True)
                
                for row_idx in summary_row_indices:
                    excel_row = row_idx + 2
                    for col in range(1, 14):  # 所有列（添加Driver Cancel后共13列）
                        cell = worksheet.cell(row=excel_row, column=col)
                        cell.fill = yellow_fill
                        cell.font = bold_font
                
                # 设置总计行样式（橙色背景+加粗）
                orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                for col in range(1, 14):  # 所有列（添加Driver Cancel后共13列）
                    cell = worksheet.cell(row=last_row, column=col)
                    cell.fill = orange_fill
                    cell.font = bold_font
                
                # 设置货币格式（$符号）
                for row in range(2, last_row + 1):
                    # 订单价格 (I列)
                    worksheet.cell(row=row, column=9).number_format = '$#,##0.00'
                    # NO SHOW (J列)
                    worksheet.cell(row=row, column=10).number_format = '$#,##0.00'
                    # Co Pay (K列)
                    worksheet.cell(row=row, column=11).number_format = '$#,##0.00'
                    # TOLL (L列)
                    worksheet.cell(row=row, column=12).number_format = '$#,##0.00'
                    # 总收入 (C列)
                    if worksheet.cell(row=row, column=3).value:
                        worksheet.cell(row=row, column=3).number_format = '$#,##0.00'
                
                # 调整列宽
                worksheet.column_dimensions['A'].width = 20  # 司机姓名
                worksheet.column_dimensions['B'].width = 10  # 订单数
                worksheet.column_dimensions['C'].width = 12  # 总收入
                worksheet.column_dimensions['D'].width = 12  # 订单ID
                worksheet.column_dimensions['E'].width = 20  # 接客时间
                worksheet.column_dimensions['F'].width = 45  # 接客地点
                worksheet.column_dimensions['G'].width = 45  # 送达地点
                worksheet.column_dimensions['H'].width = 20  # 乘客姓名
                worksheet.column_dimensions['I'].width = 15  # 订单价格
                worksheet.column_dimensions['J'].width = 12  # NO SHOW
                worksheet.column_dimensions['K'].width = 12  # Co Pay
                worksheet.column_dimensions['L'].width = 12  # TOLL
                worksheet.column_dimensions['M'].width = 15  # 状态
            
            self.log(f"✓ 账单已导出到: {filename}", "success")
            messagebox.showinfo("成功", f"账单导出成功！\n\n文件: {filename}")
            
        except ImportError:
            self.log("✗ 需要安装 pandas 和 openpyxl 库", "error")
            messagebox.showerror("错误", "缺少必要的库\n\n请安装:\npip install pandas openpyxl")
        except Exception as e:
            self.log(f"✗ 导出失败: {str(e)}", "error")
            messagebox.showerror("错误", f"导出失败:\n{e}")
    
    # ==================== 系统功能 ====================
    
    def _export_json_file(self, data, filename):
        """辅助方法：导出JSON文件（用于多线程）"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return f"✓ JSON已导出: {filename}"
        except Exception as e:
            raise Exception(f"JSON导出失败: {e}")
    
    def view_logs(self):
        """查看日志"""
        try:
            if os.path.exists(config.LOG_FILE):
                os.startfile(config.LOG_FILE)
            else:
                messagebox.showwarning("警告", "日志文件不存在")
        except Exception as e:
            messagebox.showerror("错误", f"打开日志失败: {str(e)}")
    
    def clear_output(self):
        """清空输出"""
        self.output_text.delete(1.0, tk.END)
        self.log("=" * 60)
        self.log("输出已清空", "info")
        self.log("=" * 60)
    
    def export_drivers_excel(self, drivers, filename):
        """导出司机完整数据为Excel（包含所有证件和车辆信息）"""
        try:
            import pandas as pd
            
            # 准备数据
            data = []
            for driver_data in drivers:
                # API返回的数据结构可能是:
                # 1. {driver: {...}, documents: [...], cars: [...]} - 来自get_driver_detail
                # 2. {id: ..., first_name: ..., driver: {...}, ...} - 来自get_all_drivers_with_full_details合并后
                # 需要智能识别并从正确位置提取
                
                # 优先从嵌套的driver对象获取，如果没有则从顶层获取
                if 'driver' in driver_data and isinstance(driver_data['driver'], dict):
                    driver = driver_data['driver']
                else:
                    driver = driver_data
                
                # 基本信息 - 从driver对象中提取
                driver_id = driver.get('id', '') or driver_data.get('id', '')
                first_name = driver.get('first_name', '') or driver_data.get('first_name', '')
                last_name = driver.get('last_name', '') or driver_data.get('last_name', '')
                middle_name = driver.get('middle_name', '') or driver_data.get('middle_name', '') or ''
                name = f"{first_name} {last_name}".strip() or driver.get('name', '') or driver_data.get('name', '')
                
                # 联系方式
                phone = (driver.get('phone_number', '') or driver_data.get('phone_number', '') or 
                        driver.get('phone', '') or driver_data.get('phone', '') or 
                        driver.get('mobile', '') or driver_data.get('mobile', ''))
                email = driver.get('email', '') or driver_data.get('email', '')
                
                # 地址信息
                address = (driver.get('address_street', '') or driver_data.get('address_street', '') or 
                          driver.get('address', '') or driver_data.get('address', '') or 
                          driver.get('street_address', '') or driver_data.get('street_address', ''))
                city = driver.get('address_city', '') or driver_data.get('address_city', '') or driver.get('city', '') or driver_data.get('city', '')
                state = driver.get('address_state', '') or driver_data.get('address_state', '') or driver.get('state', '') or driver_data.get('state', '')
                zip_code = (driver.get('address_zipcode', '') or driver_data.get('address_zipcode', '') or 
                           driver.get('zip_code', '') or driver_data.get('zip_code', '') or 
                           driver.get('postal_code', '') or driver_data.get('postal_code', ''))
                
                # 个人信息
                dob = driver.get('dob_date', '') or driver_data.get('dob_date', '') or driver.get('date_of_birth', '') or driver_data.get('date_of_birth', '')
                ssn = driver.get('ssn', '') or driver_data.get('ssn', '') or driver.get('social_security_number', '') or driver_data.get('social_security_number', '')
                sex = driver.get('sex', '') or driver_data.get('sex', '')
                
                # 从documents数组中提取证件信息 - 从原始driver_data中获取
                documents = driver_data.get('documents', [])
                
                # 初始化所有证件字段
                driver_license_number = driver_license_issue_date = driver_license_expiry = ''
                driver_license_state = driver_license_class = ''
                tlc_license_number = tlc_license_expiry = ''
                sentry_drug_test_number = sentry_drug_test_expiry = sentry_drug_test_status = ''
                arro_drug_test_number = arro_drug_test_expiry = arro_drug_test_status = ''
                
                # 遍历documents数组提取证件信息
                for doc in documents:
                    doc_type = doc.get('type', '')
                    
                    if doc_type == 'driver_license':
                        driver_license_number = doc.get('number', '')
                        driver_license_expiry = doc.get('expires_at', '')
                        driver_license_state = doc.get('state', '')
                        # 从options中提取issue_date和license_class
                        options = doc.get('options', [])
                        if isinstance(options, list):
                            for opt in options:
                                if opt.get('name') == 'issue_date':
                                    driver_license_issue_date = opt.get('value', '')
                                elif opt.get('name') == 'license_class':
                                    driver_license_class = opt.get('value', '')
                    
                    elif doc_type == 'tlc_license':
                        tlc_license_number = doc.get('number', '')
                        tlc_license_expiry = doc.get('expires_at', '')
                    
                    elif doc_type == 'sentry_drug_test':
                        sentry_drug_test_number = doc.get('number', '')
                        sentry_drug_test_expiry = doc.get('expires_at', '')
                        sentry_drug_test_status = doc.get('status', '')
                    
                    elif doc_type == 'arro_drug_test':
                        arro_drug_test_number = doc.get('number', '')
                        arro_drug_test_expiry = doc.get('expires_at', '')
                        arro_drug_test_status = doc.get('status', '')
                
                # 获取车辆信息 - 从driver_data中提取
                cars = driver_data.get('cars', []) or driver.get('cars', [])
                vehicle_detail = driver_data.get('vehicle_detail', {}) or driver.get('vehicle_detail', {})
                
                # 优先使用vehicle_detail，否则使用cars数组第一辆车
                if vehicle_detail and isinstance(vehicle_detail, dict):
                    vehicle = vehicle_detail
                elif cars and isinstance(cars, list) and len(cars) > 0:
                    vehicle = cars[0]
                else:
                    vehicle = {}
                
                # 初始化车辆字段
                vin = make = model = year = plate = color = vehicle_type = vehicle_state = ''
                seats = wav_seats = ''
                fhv_diamond_number = fhv_diamond_expiry = fhv_diamond_state = ''
                insurance_number = insurance_expiry = insurance_state = insurance_company = insurance_effective_date = ''
                registration_number = registration_expiry = registration_state = ''
                inspection_number = inspection_expiry = ''
                
                # 提取车辆基本信息
                if isinstance(vehicle, dict) and vehicle:
                    vin = vehicle.get('vin_number', '') or vehicle.get('vin', '')
                    make = vehicle.get('make', '')
                    model = vehicle.get('model', '')
                    year = vehicle.get('year', '')
                    plate = vehicle.get('plate_number', '') or vehicle.get('number_display', '') or vehicle.get('plate', '')
                    color = vehicle.get('color', '')
                    vehicle_type = vehicle.get('type', '')
                    vehicle_state = vehicle.get('state', '')
                    seats = vehicle.get('seats', '')
                    wav_seats = vehicle.get('wav_seats', '')
                    
                    # 从车辆的documents数组中提取证件信息
                    car_documents = vehicle.get('documents', [])
                    for doc in car_documents:
                        doc_type = doc.get('type', '')
                        
                        if doc_type == 'fhv_diamond':
                            fhv_diamond_number = doc.get('number', '')
                            fhv_diamond_expiry = doc.get('expires_at', '')
                            fhv_diamond_state = doc.get('state', '')
                        
                        elif doc_type == 'insurance_id_card':
                            insurance_number = doc.get('number', '')
                            insurance_expiry = doc.get('expires_at', '')
                            insurance_state = doc.get('state', '')
                            # 从options中提取insurance_company和effective_date
                            options = doc.get('options', [])
                            if isinstance(options, list):
                                for opt in options:
                                    if opt.get('name') == 'insurance_company':
                                        insurance_company = opt.get('value', '')
                                    elif opt.get('name') == 'effective_date':
                                        insurance_effective_date = opt.get('value', '')
                        
                        elif doc_type == 'registration':
                            registration_number = doc.get('number', '')
                            registration_expiry = doc.get('expires_at', '')
                            registration_state = doc.get('state', '')
                        
                        elif doc_type == 'nys_inspection_sticker':
                            inspection_number = doc.get('number', '')
                            inspection_expiry = doc.get('expires_at', '')
                
                # 其他信息
                status = driver.get('status', '')
                created_at = driver.get('created_at', '')
                updated_at = driver.get('updated_at', '')
                
                # 组装数据行 - 使用英文字段名
                data.append({
                    # Driver Basic Info
                    'Driver ID': driver_id,
                    'Name': name,
                    'First Name': first_name,
                    'Middle Name': middle_name,
                    'Last Name': last_name,
                    'Date of Birth': dob,
                    'SSN': ssn,
                    'Sex': sex,
                    'Email': email,
                    'Phone': phone,
                    
                    # Address
                    'Address': address,
                    'City': city,
                    'State': state,
                    'Zip Code': zip_code,
                    
                    # Driver License
                    'Driver License Number': driver_license_number,
                    'Driver License Issue Date': driver_license_issue_date,
                    'Driver License Expired Date': driver_license_expiry,
                    'Driver License State': driver_license_state,
                    'Driver License Class': driver_license_class,
                    
                    # TLC License
                    'TLC License Number': tlc_license_number,
                    'TLC License Expired Date': tlc_license_expiry,
                    
                    # Drug Tests
                    'Sentry Drug Test Number': sentry_drug_test_number,
                    'Sentry Drug Test Expired Date': sentry_drug_test_expiry,
                    'Sentry Drug Test Status': sentry_drug_test_status,
                    
                    'ARRO Drug Test Number': arro_drug_test_number,
                    'ARRO Drug Test Expired Date': arro_drug_test_expiry,
                    'ARRO Drug Test Status': arro_drug_test_status,
                    
                    # Car Basic Info
                    'VIN Number': vin,
                    'Make': make,
                    'Model': model,
                    'Year': year,
                    'Plate Number': plate,
                    'Color': color,
                    'Type': vehicle_type,
                    'Vehicle State': vehicle_state,
                    'Seats': seats,
                    'WAV Seats': wav_seats,
                    
                    # Car Documents
                    'FHV Diamond Number': fhv_diamond_number,
                    'FHV Diamond Expired Date': fhv_diamond_expiry,
                    'FHV Diamond State': fhv_diamond_state,
                    
                    'Insurance Policy Number': insurance_number,
                    'Insurance Expired Date': insurance_expiry,
                    'Insurance State': insurance_state,
                    'Insurance Company': insurance_company,
                    'Insurance Effective Date': insurance_effective_date,
                    
                    'Registration Number': registration_number,
                    'Registration Expired Date': registration_expiry,
                    'Registration State': registration_state,
                    
                    'NYS Inspection Sticker Number': inspection_number,
                    'NYS Inspection Sticker Expired Date': inspection_expiry,
                    
                    # Status
                    'Status': status,
                    'Created At': created_at,
                    'Updated At': updated_at
                })
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 导出Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='司机完整数据', index=False)
                
                # 格式化
                worksheet = writer.sheets['司机完整数据']
                
                # 设置列宽
                for idx, col in enumerate(df.columns, 1):
                    # 计算列宽
                    max_length = len(str(col))
                    for value in df.iloc[:, idx-1]:
                        try:
                            if len(str(value)) > max_length:
                                max_length = len(str(value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = adjusted_width
                
                # 冻结首行
                worksheet.freeze_panes = 'A2'
                
                # 设置表头样式
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            return f"✓ Excel已导出: {filename}\n✓ 共导出 {len(data)} 位司机的完整数据"
        except Exception as e:
            import traceback
            logger.error(f"Excel导出失败: {e}\n{traceback.format_exc()}")
            raise Exception(f"Excel导出失败: {e}")
    
    def export_schedules_excel(self, routes, filename, date):
        """导出排班数据为Excel（优化用于多线程）"""
        try:
            import pandas as pd
            
            # 准备数据
            data = []
            for route in routes:
                driver_info = route.get('driver', {}) or {}
                vehicle_info = route.get('vehicle', {}) or {}
                
                # 提取司机信息
                driver_name = driver_info.get('name', '') or f"{driver_info.get('first_name', '')} {driver_info.get('last_name', '')}".strip()
                
                data.append({
                    '路线ID': route.get('id', ''),
                    '日期': date,
                    '司机ID': driver_info.get('id', '') or route.get('driver_id', ''),
                    '司机姓名': driver_name,
                    '司机电话': driver_info.get('phone', '') or driver_info.get('mobile', ''),
                    '车辆ID': vehicle_info.get('id', '') or route.get('vehicle_id', ''),
                    '车牌号': vehicle_info.get('plate', '') or vehicle_info.get('license_plate', ''),
                    '车型': vehicle_info.get('model', '') or vehicle_info.get('make_model', ''),
                    '开工时间': route.get('start_time', '') or route.get('clock_in_time', '') or route.get('from_datetime', '') or route.get('scheduled_start', ''),
                    '收工时间': route.get('end_time', '') or route.get('clock_out_time', '') or route.get('to_datetime', '') or route.get('scheduled_end', ''),
                    '计划出发时间': route.get('scheduled_start_time', '') or route.get('planned_start', ''),
                    '实际出发时间': route.get('actual_start_time', '') or route.get('started_at', ''),
                    '计划结束时间': route.get('scheduled_end_time', '') or route.get('planned_end', ''),
                    '实际结束时间': route.get('actual_end_time', '') or route.get('ended_at', ''),
                    '起点': route.get('start_location', '') or route.get('start_address', '') or route.get('origin', ''),
                    '终点': route.get('end_location', '') or route.get('end_address', '') or route.get('destination', ''),
                    '总里程': route.get('total_distance', '') or route.get('distance', ''),
                    '总时长': route.get('total_duration', '') or route.get('duration', ''),
                    '订单数': route.get('rides_count', '') or route.get('total_rides', '') or route.get('ride_count', ''),
                    '状态': route.get('status', ''),
                    '备注': route.get('notes', '') or route.get('comment', '')
                })
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 导出Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='排班数据', index=False)
                
                # 格式化
                worksheet = writer.sheets['排班数据']
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
            
            return f"✓ Excel已导出: {filename}"
        except Exception as e:
            raise Exception(f"Excel导出失败: {e}")
    
    def export_orders_excel(self, rides, filename, date):
        """导出订单数据为Excel（优化用于多线程）"""
        try:
            import pandas as pd
            
            # 准备数据
            data = []
            for ride in rides:
                driver_info = ride.get('driver', {}) or {}
                passenger_info = ride.get('passenger', {}) or ride.get('customer', {}) or {}
                
                # 提取司机信息
                driver_name = driver_info.get('name', '') or f"{driver_info.get('first_name', '')} {driver_info.get('last_name', '')}".strip() or f"{ride.get('driver_first_name', '')} {ride.get('driver_last_name', '')}".strip()
                
                # 提取乘客信息
                passenger_name = passenger_info.get('name', '') or f"{passenger_info.get('first_name', '')} {passenger_info.get('last_name', '')}".strip() or f"{ride.get('first_name', '')} {ride.get('last_name', '')}".strip() or ride.get('customer_name', '')
                
                data.append({
                    '订单ID': ride.get('id', ''),
                    '日期': date,
                    '司机ID': driver_info.get('id', '') or ride.get('driver_id', ''),
                    '司机姓名': driver_name,
                    '司机电话': driver_info.get('phone', '') or driver_info.get('mobile', ''),
                    '乘客ID': passenger_info.get('id', '') or ride.get('customer_id', ''),
                    '乘客姓名': passenger_name,
                    '乘客电话': passenger_info.get('phone', '') or passenger_info.get('mobile', '') or ride.get('customer_phone', ''),
                    '接客地址': ride.get('pickup_address', '') or ride.get('start_address', '') or ride.get('origin_address', ''),
                    '送达地址': ride.get('dropoff_address', '') or ride.get('destination_address', '') or ride.get('dest_address', ''),
                    '计划接客时间': ride.get('pickup_at', '') or ride.get('schedule_time', '') or ride.get('scheduled_pickup', ''),
                    '实际接客时间': ride.get('actual_pickup_time', '') or ride.get('pickup_time', '') or ride.get('picked_up_at', ''),
                    '计划送达时间': ride.get('scheduled_dropoff_time', '') or ride.get('scheduled_dropoff', ''),
                    '实际送达时间': ride.get('actual_dropoff_time', '') or ride.get('dropoff_time', '') or ride.get('dropped_off_at', ''),
                    '订单价格': float(ride.get('order_price', 0) or ride.get('price', 0) or ride.get('base_price', 0) or ride.get('vendor_amount', 0) or 0),
                    'Co-Pay': float(ride.get('co_pay', 0) or ride.get('copay', 0) or 0),
                    'Toll费': float(ride.get('toll_fee', 0) or ride.get('toll', 0) or ride.get('tolls', 0) or 0),
                    '小费': float(ride.get('tip', 0) or ride.get('gratuity', 0) or 0),
                    '总金额': float(ride.get('total_amount', 0) or ride.get('total', 0) or ride.get('vendor_amount', 0) or 0),
                    '距离(英里)': float(ride.get('distance', 0) or 0),
                    '行驶时长': ride.get('duration', '') or ride.get('drive_time', ''),
                    '状态': ride.get('status', ''),
                    '支付方式': ride.get('payment_method', '') or ride.get('payment_type', ''),
                    '订单类型': ride.get('ride_type', '') or ride.get('service_type', ''),
                    '备注': ride.get('notes', '') or ride.get('comment', '') or ride.get('description', '')
                })
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 导出Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='订单数据', index=False)
                
                # 格式化
                worksheet = writer.sheets['订单数据']
                from openpyxl.styles import numbers
                
                # 自动调整所有列宽
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # 价格列格式化为货币（查找包含价格、金额、费用的列）
                for col_idx, col_name in enumerate(df.columns, start=1):
                    if any(keyword in str(col_name) for keyword in ['价格', '金额', 'Pay', 'Toll', '小费']):
                        for row in range(2, len(rides) + 2):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.number_format = '$#,##0.00'
            
            return f"✓ Excel已导出: {filename}"
        except Exception as e:
            raise Exception(f"Excel导出失败: {e}")
    
    def show_about(self):
        """显示关于"""
        about_text = """
RPA数据爬取工具 v1.0

专注于数据采集和导出功能

功能：
• 爬取司机、排班、订单数据
• 生成账单
• 导出JSON/Excel
• 数据管理

技术支持：请联系管理员
        """
        messagebox.showinfo("关于", about_text)


def main():
    try:
        root = tk.Tk()
        app = DataScraperGUI(root)
        root.mainloop()
    except Exception as e:
        import traceback
        error_msg = f"启动失败:\n{str(e)}\n\n{traceback.format_exc()}"
        print(error_msg)
        # 尝试显示错误对话框
        try:
            import tkinter.messagebox as mb
            root = tk.Tk()
            root.withdraw()
            mb.showerror("启动错误", error_msg)
        except:
            pass


if __name__ == "__main__":
    main()
